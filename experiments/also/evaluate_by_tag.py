#!/usr/bin/env python3
"""
Evaluate experiments from outputs directory and calculate metrics.

This script reads experiment results from the outputs directory and calculates
metrics. Each experiment folder is treated as a separate experiment.

Usage:
    python evaluate_by_tag.py --list-tags  # List all experiments
    python evaluate_by_tag.py --tag bandit_adversarial_p1_20251204_181158
    python evaluate_by_tag.py --tags exp1 exp2 exp3  # Compare multiple experiments
    python evaluate_by_tag.py --tag exp1 --eval-set hard  # Evaluate only hard subset
"""

import argparse
import json
import math
import re
from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import yaml
from loguru import logger
from rich.console import Console
from rich.table import Table
from rich.panel import Panel

# Load environment variables
from dotenv import load_dotenv
load_dotenv()

console = Console()

# Default outputs directory
OUTPUTS_DIR = Path(__file__).parent / "outputs"

# Path to env_ids.txt file
ENV_IDS_FILE = Path(__file__).parent.parent.parent / "data" / "env_ids.txt"

EVALSET_DISPLAY_NAMES = {
    "all": "SOTOPIA",
    "hard": "SOTOPIA-Hard",
}

SUMMARY_METRICS = [
    ("goal", "Goal"),
    ("relationship", "Relation"),
    ("knowledge", "Knowledge"),
    ("overall", "Overall"),
]


def load_env_ids_by_dataset(eval_set: str = "all") -> set[str] | None:
    """
    Load environment IDs from env_ids.txt based on the selected dataset.

    Args:
        eval_set: Either "all" or "hard" to select the dataset

    Returns:
        Set of environment IDs for the selected dataset, or None if no filtering needed
    """
    if eval_set == "all":
        # No filtering needed for "all" - return None to indicate no filtering
        return None

    if not ENV_IDS_FILE.exists():
        logger.warning(f"env_ids.txt not found at {ENV_IDS_FILE}")
        return None

    try:
        with open(ENV_IDS_FILE, "r") as f:
            content = f.read()

        # Parse the file to extract environment IDs for each dataset
        hard_ids: set[str] = set()
        all_ids: set[str] = set()

        # Find SOTOPIA-HARD section
        hard_match = re.search(r"SOTOPIA-HARD:\s*\n?(.*?)(?=\n\s*SOTOPIA-|$)", content, re.DOTALL)
        if hard_match:
            hard_line = hard_match.group(1).strip()
            # Extract IDs from quoted strings
            hard_ids = set(re.findall(r'"([A-Z0-9]+)"', hard_line))
            logger.debug(f"Loaded {len(hard_ids)} hard environment IDs")

        # Find SOTOPIA-ALL section
        all_match = re.search(r"SOTOPIA-ALL:\s*\n?(.*?)$", content, re.DOTALL)
        if all_match:
            all_line = all_match.group(1).strip()
            all_ids = set(re.findall(r'"([A-Z0-9]+)"', all_line))
            logger.debug(f"Loaded {len(all_ids)} all environment IDs")

        if eval_set == "hard":
            if not hard_ids:
                logger.warning("No hard environment IDs found in env_ids.txt")
                return None
            return hard_ids
        else:
            # For any other value, return None (no filtering)
            return None

    except Exception as e:
        logger.error(f"Error loading env_ids.txt: {e}")
        return None

# Evaluation dimensions
DIMENSIONS = [
    "believability",
    "relationship",
    "knowledge",
    "secret",
    "social_rules",
    "financial_and_material_benefits",
    "goal",
]

# Known bandit types (order matters: longer names first to match correctly)
KNOWN_BANDIT_TYPES = [
    "prompt_breeder",
    "neural_ucb",
    "neural_evolution",
    "adversarial",
    "exp3",
    "none",
]


def parse_bandit_type_from_name(experiment_name: str) -> str:
    """Parse bandit type from experiment name.

    Experiment names follow the pattern: bandit_{type}_{optimize}_{subset}_...
    Where type can be: adversarial, exp3, neural_ucb, prompt_breeder, none
    """
    # Remove 'bandit_' prefix if present
    name = experiment_name
    if name.startswith("bandit_"):
        name = name[7:]  # len("bandit_") = 7

    # Try to match known bandit types (longer names first)
    for btype in KNOWN_BANDIT_TYPES:
        if name.startswith(btype + "_") or name.startswith(btype):
            return btype

    return "unknown"


def parse_optimize_mode_from_name(experiment_name: str) -> str:
    """Parse optimize mode from experiment name.

    Experiment names follow the pattern: bandit_{type}_{optimize}_{subset}_...
    Where optimize can be: p1, p2, both, none
    """
    # Known optimize modes
    optimize_modes = ["both", "p1", "p2", "none"]

    # Remove 'bandit_' prefix if present
    name = experiment_name
    if name.startswith("bandit_"):
        name = name[7:]

    # Remove bandit type prefix
    for btype in KNOWN_BANDIT_TYPES:
        if name.startswith(btype + "_"):
            name = name[len(btype) + 1:]
            break

    # Now name should start with optimize mode
    for mode in optimize_modes:
        if name.startswith(mode + "_") or name.startswith(mode):
            return mode

    return "unknown"


def get_db_final_rewards(tag: str, env_ids_filter: set[str] | None = None) -> dict[str, Any] | None:
    """
    Get final rewards from database for a given tag.

    Args:
        tag: The experiment tag to query
        env_ids_filter: Optional set of environment IDs to filter episodes by.
                        If None, all episodes with the tag are included.

    Returns:
        Dict with p1_final_mean, p1_final_std, p1_final_se, p2_final_mean, p2_final_std, p2_final_se,
        p1_breakdown (mean, std, & se for each dimension), p2_breakdown,
        plus p1_rewards_raw and p2_rewards_raw lists for CI calculation,
        plus data_quality metrics (tuple_count, float_count), or None if not found
    """
    try:
        from sotopia.database.logs import EpisodeLog
    except ImportError:
        logger.warning("Could not import EpisodeLog, skipping database query")
        return None

    try:
        episodes = list(EpisodeLog.find(EpisodeLog.tag == tag).all())
    except Exception as e:
        logger.warning(f"Failed to query database for tag {tag}: {e}")
        return None

    if not episodes:
        logger.debug(f"No episodes found in database for tag: {tag}")
        return None

    # Filter episodes by environment IDs if specified
    if env_ids_filter is not None:
        original_count = len(episodes)
        episodes = [ep for ep in episodes if ep.environment in env_ids_filter]
        filtered_count = len(episodes)
        if filtered_count < original_count:
            logger.debug(f"Filtered {original_count} -> {filtered_count} episodes by env_ids")
        if not episodes:
            logger.debug(f"No episodes remaining after env_id filtering for tag: {tag}")
            return None

    p1_rewards = []
    p2_rewards = []
    p1_breakdowns: list[dict[str, float]] = []
    p2_breakdowns: list[dict[str, float]] = []

    # Data quality tracking
    p1_tuple_count = 0
    p1_float_count = 0
    p2_tuple_count = 0
    p2_float_count = 0

    for episode in episodes:
        if len(episode.rewards) >= 2:
            # Parse reward - can be float or tuple(float, dict)
            r1 = episode.rewards[0]
            r2 = episode.rewards[1]

            if isinstance(r1, tuple) and len(r1) == 2:
                p1_tuple_count += 1
                p1_rewards.append(float(r1[0]))
                if r1[1]:
                    p1_breakdowns.append(r1[1])
            else:
                p1_float_count += 1
                p1_rewards.append(float(r1))

            if isinstance(r2, tuple) and len(r2) == 2:
                p2_tuple_count += 1
                p2_rewards.append(float(r2[0]))
                if r2[1]:
                    p2_breakdowns.append(r2[1])
            else:
                p2_float_count += 1
                p2_rewards.append(float(r2))

    if not p1_rewards:
        return None

    # Calculate standard error helper (inline to avoid forward reference)
    def _calc_se(values: list[float]) -> float:
        if len(values) <= 1:
            return 0.0
        return float(np.std(values)) / math.sqrt(len(values))

    result: dict[str, Any] = {
        "episode_count": len(episodes),
        "p1_final_mean": float(np.mean(p1_rewards)),
        "p1_final_std": float(np.std(p1_rewards)),
        "p1_final_se": _calc_se(p1_rewards),
        "p2_final_mean": float(np.mean(p2_rewards)) if p2_rewards else 0.0,
        "p2_final_std": float(np.std(p2_rewards)) if p2_rewards else 0.0,
        "p2_final_se": _calc_se(p2_rewards) if p2_rewards else 0.0,
        # Include raw values for CI calculation
        "p1_rewards_raw": p1_rewards,
        "p2_rewards_raw": p2_rewards,
        # Data quality metrics
        "p1_tuple_count": p1_tuple_count,
        "p1_float_count": p1_float_count,
        "p2_tuple_count": p2_tuple_count,
        "p2_float_count": p2_float_count,
        "data_quality_status": "healthy" if (p1_float_count == 0 and p2_float_count == 0) else "degraded",
    }

    # Aggregate breakdowns with mean, std, and se for each dimension
    if p1_breakdowns:
        keys = list(p1_breakdowns[0].keys())
        result["p1_breakdown"] = {}
        result["p1_breakdown_std"] = {}
        result["p1_breakdown_se"] = {}
        result["p1_breakdown_raw"] = {}  # Store raw values for CI calculation
        for k in keys:
            values = [b.get(k, 0) for b in p1_breakdowns]
            result["p1_breakdown"][k] = float(np.mean(values))
            result["p1_breakdown_std"][k] = float(np.std(values))
            result["p1_breakdown_se"][k] = _calc_se(values)
            result["p1_breakdown_raw"][k] = values  # Raw values

    if p2_breakdowns:
        keys = list(p2_breakdowns[0].keys())
        result["p2_breakdown"] = {}
        result["p2_breakdown_std"] = {}
        result["p2_breakdown_se"] = {}
        result["p2_breakdown_raw"] = {}  # Store raw values for CI calculation
        for k in keys:
            values = [b.get(k, 0) for b in p2_breakdowns]
            result["p2_breakdown"][k] = float(np.mean(values))
            result["p2_breakdown_std"][k] = float(np.std(values))
            result["p2_breakdown_se"][k] = _calc_se(values)
            result["p2_breakdown_raw"][k] = values  # Raw values

    return result


def get_local_final_rewards(experiment_dir: Path, env_ids_filter: set[str] | None = None) -> dict[str, Any] | None:
    """
    Get final rewards from local results.json file.

    This is a fallback when database is not available or --push-to-db was not used.
    Reads the 'final_rewards' field that contains dimension breakdowns.

    Args:
        experiment_dir: Path to the experiment output directory
        env_ids_filter: Optional set of environment IDs to filter results by.
                        If None, all results are included.

    Returns:
        Dict with same structure as get_db_final_rewards, or None if not found
    """
    results_file = experiment_dir / "results.json"
    if not results_file.exists():
        return None

    try:
        with open(results_file) as f:
            data = json.load(f)
    except json.JSONDecodeError:
        logger.warning(f"Invalid JSON in {results_file}")
        return None

    results = data.get("results", [])
    if not results:
        return None

    # Filter results by environment IDs if specified
    if env_ids_filter is not None:
        original_count = len(results)
        filtered_results = []
        for result in results:
            # Try to get env_id from various possible locations in the result
            summary = result.get("summary", {})
            env_id = (
                result.get("env_id")
                or result.get("environment_id")
                or summary.get("env_id")
                or summary.get("environment_id")
                or summary.get("scenario_id")
            )
            if env_id and env_id in env_ids_filter:
                filtered_results.append(result)
        results = filtered_results
        filtered_count = len(results)
        if filtered_count < original_count:
            logger.debug(f"Filtered {original_count} -> {filtered_count} results by env_ids")
        if not results:
            logger.debug(f"No results remaining after env_id filtering in {experiment_dir}")
            return None

    p1_rewards = []
    p2_rewards = []
    p1_breakdowns: list[dict[str, float]] = []
    p2_breakdowns: list[dict[str, float]] = []

    # Data quality tracking
    p1_tuple_count = 0
    p1_float_count = 0
    p2_tuple_count = 0
    p2_float_count = 0

    for result in results:
        if not result.get("success"):
            continue
        summary = result.get("summary", {})
        final_rewards = summary.get("final_rewards")

        if final_rewards:
            # New format with dimension breakdowns
            p1_data = final_rewards.get("p1")
            p2_data = final_rewards.get("p2")

            if p1_data:
                p1_rewards.append(float(p1_data.get("overall", 0)))
                if p1_data.get("breakdown"):
                    p1_tuple_count += 1
                    p1_breakdowns.append(p1_data["breakdown"])
                else:
                    p1_float_count += 1

            if p2_data:
                p2_rewards.append(float(p2_data.get("overall", 0)))
                if p2_data.get("breakdown"):
                    p2_tuple_count += 1
                    p2_breakdowns.append(p2_data["breakdown"])
                else:
                    p2_float_count += 1
        else:
            # Fallback: use last turn rewards (no dimension breakdown available)
            turn_rewards = summary.get("turn_rewards", [])
            if turn_rewards:
                last_turn = turn_rewards[-1]
                p1_reward = last_turn.get("p1_reward")
                p2_reward = last_turn.get("p2_reward")
                if p1_reward is not None:
                    p1_float_count += 1
                    p1_rewards.append(float(p1_reward))
                if p2_reward is not None:
                    p2_float_count += 1
                    p2_rewards.append(float(p2_reward))

    if not p1_rewards:
        return None

    def _calc_se(values: list[float]) -> float:
        if len(values) <= 1:
            return 0.0
        return float(np.std(values)) / math.sqrt(len(values))

    result_dict: dict[str, Any] = {
        "episode_count": len(p1_rewards),
        "p1_final_mean": float(np.mean(p1_rewards)),
        "p1_final_std": float(np.std(p1_rewards)),
        "p1_final_se": _calc_se(p1_rewards),
        "p2_final_mean": float(np.mean(p2_rewards)) if p2_rewards else 0.0,
        "p2_final_std": float(np.std(p2_rewards)) if p2_rewards else 0.0,
        "p2_final_se": _calc_se(p2_rewards) if p2_rewards else 0.0,
        "p1_rewards_raw": p1_rewards,
        "p2_rewards_raw": p2_rewards,
        "p1_tuple_count": p1_tuple_count,
        "p1_float_count": p1_float_count,
        "p2_tuple_count": p2_tuple_count,
        "p2_float_count": p2_float_count,
        "data_quality_status": "healthy" if (p1_float_count == 0 and p2_float_count == 0) else "degraded",
        "source": "local_file",  # Mark that this came from local file, not database
    }

    # Aggregate breakdowns
    if p1_breakdowns:
        keys = list(p1_breakdowns[0].keys())
        result_dict["p1_breakdown"] = {}
        result_dict["p1_breakdown_std"] = {}
        result_dict["p1_breakdown_se"] = {}
        result_dict["p1_breakdown_raw"] = {}
        for k in keys:
            values = [b.get(k, 0) for b in p1_breakdowns]
            result_dict["p1_breakdown"][k] = float(np.mean(values))
            result_dict["p1_breakdown_std"][k] = float(np.std(values))
            result_dict["p1_breakdown_se"][k] = _calc_se(values)
            result_dict["p1_breakdown_raw"][k] = values

    if p2_breakdowns:
        keys = list(p2_breakdowns[0].keys())
        result_dict["p2_breakdown"] = {}
        result_dict["p2_breakdown_std"] = {}
        result_dict["p2_breakdown_se"] = {}
        result_dict["p2_breakdown_raw"] = {}
        for k in keys:
            values = [b.get(k, 0) for b in p2_breakdowns]
            result_dict["p2_breakdown"][k] = float(np.mean(values))
            result_dict["p2_breakdown_std"][k] = float(np.std(values))
            result_dict["p2_breakdown_se"][k] = _calc_se(values)
            result_dict["p2_breakdown_raw"][k] = values

    return result_dict


def calculate_standard_error(values: list[float]) -> float:
    """
    Calculate standard error for a list of values.

    Standard Error (SE) = standard_deviation / sqrt(N)

    Args:
        values: List of numeric values

    Returns:
        Standard error value. Returns 0.0 for edge cases (N=0 or N=1).
    """
    if not values:
        return 0.0

    n = len(values)
    if n <= 1:
        # With N=0 or N=1, standard error is undefined or 0
        return 0.0

    # Calculate standard deviation (using population std to be consistent with existing code)
    std = float(np.std(values))

    # Calculate standard error: SE = std / sqrt(N)
    se = std / math.sqrt(n)

    return se


def calculate_confidence_interval(values: list[float]) -> tuple[float, float]:
    """
    Calculate 95% confidence interval for a list of values.

    Returns:
        Tuple of (mean, margin_of_error)

    Implementation follows the same approach as benchmark.py get_avg_reward function.
    """
    if not values:
        return (0.0, 0.0)

    if len(values) == 1:
        return (float(values[0]), 0.0)

    # Calculate mean
    mean = float(np.mean(values))

    # Calculate variance
    variance = float(np.var(values, ddof=1))  # Using n-1 for sample variance

    # Calculate standard error of the mean (SEM)
    sem = math.sqrt(variance / len(values))

    # Use t-distribution to calculate margin of error
    confidence_level = 0.95
    df = len(values) - 1

    # Sample from t-distribution to get critical value
    t_samples = np.random.standard_t(df=df, size=1000000)
    t_value = np.percentile(t_samples, 100 * (1 - (1 - confidence_level) / 2))

    # Calculate margin of error
    margin = t_value * sem

    return (mean, float(margin))


def load_experiment_results(experiment_name: str, outputs_dir: Path = OUTPUTS_DIR) -> dict[str, Any] | None:
    """
    Load experiment results from the outputs directory.
    Supports both results.json (old format) and batch_results.json (new format).
    """
    experiment_path = outputs_dir / experiment_name
    
    if not experiment_path.exists():
        logger.error(f"Experiment folder not found: {experiment_path}")
        return None

    results_file = experiment_path / "results.json"
    batch_results_file = experiment_path / "batch_results.json"

    # Try loading results.json (Old format)
    if results_file.exists():
        try:
            with open(results_file) as f:
                return json.load(f)
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON in {results_file}: {e}")
            return None
            
    # Try loading batch_results.json (New format)
    elif batch_results_file.exists():
        try:
            with open(batch_results_file) as f:
                data = json.load(f)
                
            # Adapt new list format to old dictionary format expected by metrics calculation
            if isinstance(data, list):
                adapted_data = {
                    "results": [{"success": True, "summary": item} for item in data],
                    "total_episodes": len(data),
                    "success_count": len(data),
                }
                # Extract metadata from first item if available
                if data:
                    adapted_data["bandit_type"] = data[0].get("bandit_type", "unknown")
                    adapted_data["optimize_mode"] = data[0].get("optimize_mode", "unknown")
                    
                    # Calculate aggregated stats for display in list_experiments
                    p1_rewards = [item.get("p1_avg_reward", 0) for item in data]
                    adapted_data["p1_avg_reward"] = sum(p1_rewards) / len(p1_rewards)
                    
                    total_turns = sum(item.get("total_turns", 0) for item in data)
                    adapted_data["total_turns"] = total_turns
                    
                return adapted_data
            return data # Should not happen based on new format, but fallback
            
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON in {batch_results_file}: {e}")
            return None
            
    else:
        logger.warning(f"No results file (results.json or batch_results.json) in: {experiment_name}")
        return None


def calculate_metrics_for_experiment(
    experiment_name: str,
    outputs_dir: Path = OUTPUTS_DIR,
    env_ids_filter: set[str] | None = None
) -> dict[str, Any]:
    """Calculate metrics for an experiment from results.json or database.

    Args:
        experiment_name: Name of the experiment to evaluate
        outputs_dir: Directory containing experiment outputs
        env_ids_filter: Optional set of environment IDs to filter episodes by.
                        If None, all episodes are included.
    """
    data = load_experiment_results(experiment_name, outputs_dir)

    if data is None:
        # Fallback: try to load from database directly
        logger.info(f"Local files not found for {experiment_name}, trying database...")
        db_rewards = get_db_final_rewards(experiment_name, env_ids_filter=env_ids_filter)
        if db_rewards:
            console.print(f"[green]✓[/] Found data in database for '{experiment_name}'")
            return _calculate_metrics_from_db(experiment_name, db_rewards)
        else:
            return {"tag": experiment_name, "error": "Results not found in local files or database"}

    # Check if this is a batch experiment (has "results" array) or single experiment
    is_batch = "results" in data and isinstance(data.get("results"), list)

    if is_batch:
        return _calculate_batch_metrics(experiment_name, data, env_ids_filter=env_ids_filter)
    else:
        return _calculate_single_metrics(experiment_name, data)


def _calculate_metrics_from_db(experiment_name: str, db_rewards: dict[str, Any]) -> dict[str, Any]:
    """Calculate metrics directly from database rewards when local files are unavailable."""
    metrics: dict[str, Any] = {
        "tag": experiment_name,
        "experiment_type": "batch",
        "rewards_source": "database",
        "subset": "unknown",
        "optimize_mode": parse_optimize_mode_from_name(experiment_name),
        "bandit_type": parse_bandit_type_from_name(experiment_name),
        "total_episodes": db_rewards.get("episode_count", 0),
        "success_count": db_rewards.get("episode_count", 0),
        "total_turns": 0,
        "avg_turns_per_episode": 0,
    }

    # Final rewards from database
    metrics["db_episode_count"] = db_rewards.get("episode_count", 0)
    metrics["p1_final_mean"] = db_rewards.get("p1_final_mean", 0)
    metrics["p1_final_std"] = db_rewards.get("p1_final_std", 0)
    metrics["p1_final_se"] = db_rewards.get("p1_final_se", 0)
    metrics["p2_final_mean"] = db_rewards.get("p2_final_mean", 0)
    metrics["p2_final_std"] = db_rewards.get("p2_final_std", 0)
    metrics["p2_final_se"] = db_rewards.get("p2_final_se", 0)

    # Use final rewards as overall rewards since we don't have turn data
    metrics["p1_overall_mean"] = metrics["p1_final_mean"]
    metrics["p1_overall_std"] = metrics["p1_final_std"]
    metrics["p1_overall_se"] = metrics["p1_final_se"]
    metrics["p2_overall_mean"] = metrics["p2_final_mean"]
    metrics["p2_overall_std"] = metrics["p2_final_std"]
    metrics["p2_overall_se"] = metrics["p2_final_se"]
    metrics["avg_overall"] = (metrics["p1_final_mean"] + metrics["p2_final_mean"]) / 2

    # Raw rewards for CI calculation
    metrics["p1_final_rewards_raw"] = db_rewards.get("p1_rewards_raw", [])
    metrics["p2_final_rewards_raw"] = db_rewards.get("p2_rewards_raw", [])

    # Data quality metrics
    metrics["p1_tuple_count"] = db_rewards.get("p1_tuple_count", 0)
    metrics["p1_float_count"] = db_rewards.get("p1_float_count", 0)
    metrics["p2_tuple_count"] = db_rewards.get("p2_tuple_count", 0)
    metrics["p2_float_count"] = db_rewards.get("p2_float_count", 0)
    metrics["data_quality_status"] = db_rewards.get("data_quality_status", "unknown")

    # Dimension breakdowns
    if "p1_breakdown" in db_rewards:
        metrics["p1_breakdown"] = db_rewards["p1_breakdown"]
        metrics["p1_goal"] = db_rewards["p1_breakdown"].get("goal", 0)
    if "p1_breakdown_std" in db_rewards:
        metrics["p1_breakdown_std"] = db_rewards["p1_breakdown_std"]
    if "p1_breakdown_se" in db_rewards:
        metrics["p1_breakdown_se"] = db_rewards["p1_breakdown_se"]
    if "p1_breakdown_raw" in db_rewards:
        metrics["p1_breakdown_raw"] = db_rewards["p1_breakdown_raw"]
    if "p2_breakdown" in db_rewards:
        metrics["p2_breakdown"] = db_rewards["p2_breakdown"]
        metrics["p2_goal"] = db_rewards["p2_breakdown"].get("goal", 0)
    if "p2_breakdown_std" in db_rewards:
        metrics["p2_breakdown_std"] = db_rewards["p2_breakdown_std"]
    if "p2_breakdown_se" in db_rewards:
        metrics["p2_breakdown_se"] = db_rewards["p2_breakdown_se"]
    if "p2_breakdown_raw" in db_rewards:
        metrics["p2_breakdown_raw"] = db_rewards["p2_breakdown_raw"]

    return metrics


def _calculate_batch_metrics(
    experiment_name: str,
    data: dict[str, Any],
    env_ids_filter: set[str] | None = None
) -> dict[str, Any]:
    """Calculate metrics for a batch experiment (multiple scenarios).

    Args:
        experiment_name: Name of the experiment
        data: Loaded experiment data
        env_ids_filter: Optional set of environment IDs to filter episodes by.
                        If None, all episodes are included.
    """
    results = data.get("results", [])
    successful_results = [r for r in results if r.get("success") and r.get("summary")]

    if not successful_results:
        return {"tag": experiment_name, "error": "No successful results in batch"}

    # Collect all P1 and P2 average rewards across scenarios
    p1_avg_rewards = []
    p2_avg_rewards = []
    # Also collect final rewards (last turn of each episode)
    p1_final_rewards = []
    p2_final_rewards = []
    total_turns = 0
    bandit_type = "unknown"

    for result in successful_results:
        summary = result.get("summary", {})
        p1_avg = summary.get("p1_avg_reward")
        p2_avg = summary.get("p2_avg_reward")

        if p1_avg is not None:
            p1_avg_rewards.append(p1_avg)
        if p2_avg is not None:
            p2_avg_rewards.append(p2_avg)

        # Extract final rewards from turn_rewards (last turn)
        turn_rewards = summary.get("turn_rewards", [])
        if turn_rewards:
            last_turn = turn_rewards[-1]
            p1_final = last_turn.get("p1_reward")
            p2_final = last_turn.get("p2_reward")
            if p1_final is not None:
                p1_final_rewards.append(p1_final)
            if p2_final is not None:
                p2_final_rewards.append(p2_final)

        total_turns += summary.get("total_turns", 0)
        if bandit_type == "unknown":
            bandit_type = summary.get("bandit_type", "unknown")

    # Try to load config.yaml to get eta
    config_path = OUTPUTS_DIR / experiment_name / "config.yaml"
    eta = "N/A"
    if config_path.exists():
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                config = yaml.safe_load(f)
                if config:
                    eta = config.get("bandit", {}).get("eta", "N/A")
        except Exception as e:
            logger.warning(f"Failed to read config.yaml for {experiment_name}: {e}")

    metrics: dict[str, Any] = {
        "tag": experiment_name,
        "eta": eta,
        "experiment_type": "batch",
        "subset": data.get("subset", "unknown"),
        "optimize_mode": data.get("optimize_mode", "unknown"),
        "bandit_type": bandit_type,
        "total_episodes": data.get("total_episodes", len(results)),
        "success_count": data.get("success_count", len(successful_results)),
        "total_turns": total_turns,
        "avg_turns_per_episode": total_turns / len(successful_results) if successful_results else 0,
    }

    # Aggregate rewards
    metrics["p1_overall_mean"] = float(np.mean(p1_avg_rewards)) if p1_avg_rewards else 0.0
    metrics["p1_overall_std"] = float(np.std(p1_avg_rewards)) if p1_avg_rewards else 0.0
    metrics["p1_overall_se"] = calculate_standard_error(p1_avg_rewards)
    metrics["p2_overall_mean"] = float(np.mean(p2_avg_rewards)) if p2_avg_rewards else 0.0
    metrics["p2_overall_std"] = float(np.std(p2_avg_rewards)) if p2_avg_rewards else 0.0
    metrics["p2_overall_se"] = calculate_standard_error(p2_avg_rewards)
    metrics["avg_overall"] = float(np.mean(p1_avg_rewards + p2_avg_rewards)) if (p1_avg_rewards + p2_avg_rewards) else 0.0
    metrics["avg_overall_se"] = calculate_standard_error(p1_avg_rewards + p2_avg_rewards)

    # Use pre-calculated averages from data if available
    if "avg_p1_reward" in data:
        metrics["p1_overall_mean"] = data["avg_p1_reward"]
    if "avg_p2_reward" in data:
        metrics["p2_overall_mean"] = data["avg_p2_reward"]

    # Duration info
    metrics["total_duration_seconds"] = data.get("total_duration_seconds", 0)
    metrics["avg_duration_per_episode"] = data.get("avg_duration_per_episode", 0)

    # Get final rewards from database (preferred source), fallback to local file
    db_rewards = get_db_final_rewards(experiment_name, env_ids_filter=env_ids_filter)
    rewards_source = "database"
    if db_rewards is None:
        # Warn user that database has no data
        logger.warning(
            f"[{experiment_name}] No data found in database. "
            "Did you run with --push-to-db? Falling back to local file..."
        )
        console.print(
            f"[yellow]⚠ Warning:[/] No database data for '{experiment_name}'. "
            "Consider using --push-to-db when running experiments."
        )
        # Fallback: try to read from local results.json with final_rewards field
        db_rewards = get_local_final_rewards(OUTPUTS_DIR / experiment_name, env_ids_filter=env_ids_filter)
        if db_rewards:
            rewards_source = "local_file"
            logger.info(f"Using local file final_rewards for {experiment_name}")
        else:
            rewards_source = "turn_rewards_fallback"
            logger.warning(
                f"[{experiment_name}] Local file also missing final_rewards. "
                "Dimension breakdowns (goal, etc.) will not be available."
            )
            console.print(
                f"[yellow]⚠ Warning:[/] No dimension breakdowns available for '{experiment_name}'. "
                "Re-run experiment to get goal and other dimension scores."
            )

    metrics["rewards_source"] = rewards_source
    if db_rewards:
        metrics["db_episode_count"] = db_rewards.get("episode_count", 0)
        metrics["p1_final_mean"] = db_rewards.get("p1_final_mean", 0)
        metrics["p1_final_std"] = db_rewards.get("p1_final_std", 0)
        metrics["p1_final_se"] = db_rewards.get("p1_final_se", 0)
        metrics["p2_final_mean"] = db_rewards.get("p2_final_mean", 0)
        metrics["p2_final_std"] = db_rewards.get("p2_final_std", 0)
        metrics["p2_final_se"] = db_rewards.get("p2_final_se", 0)
        # Store raw rewards for CI calculation
        metrics["p1_final_rewards_raw"] = db_rewards.get("p1_rewards_raw", [])
        metrics["p2_final_rewards_raw"] = db_rewards.get("p2_rewards_raw", [])
        # Data quality metrics
        metrics["p1_tuple_count"] = db_rewards.get("p1_tuple_count", 0)
        metrics["p1_float_count"] = db_rewards.get("p1_float_count", 0)
        metrics["p2_tuple_count"] = db_rewards.get("p2_tuple_count", 0)
        metrics["p2_float_count"] = db_rewards.get("p2_float_count", 0)
        metrics["data_quality_status"] = db_rewards.get("data_quality_status", "unknown")
        # Extract goal dimension for display
        if "p1_breakdown" in db_rewards:
            metrics["p1_breakdown"] = db_rewards["p1_breakdown"]
            metrics["p1_goal"] = db_rewards["p1_breakdown"].get("goal", 0)
        if "p1_breakdown_std" in db_rewards:
            metrics["p1_breakdown_std"] = db_rewards["p1_breakdown_std"]
        if "p1_breakdown_se" in db_rewards:
            metrics["p1_breakdown_se"] = db_rewards["p1_breakdown_se"]
        if "p1_breakdown_raw" in db_rewards:
            metrics["p1_breakdown_raw"] = db_rewards["p1_breakdown_raw"]
        if "p2_breakdown" in db_rewards:
            metrics["p2_breakdown"] = db_rewards["p2_breakdown"]
            metrics["p2_goal"] = db_rewards["p2_breakdown"].get("goal", 0)
        if "p2_breakdown_std" in db_rewards:
            metrics["p2_breakdown_std"] = db_rewards["p2_breakdown_std"]
        if "p2_breakdown_se" in db_rewards:
            metrics["p2_breakdown_se"] = db_rewards["p2_breakdown_se"]
        if "p2_breakdown_raw" in db_rewards:
            metrics["p2_breakdown_raw"] = db_rewards["p2_breakdown_raw"]
    elif p1_final_rewards:
        # Fallback: use final rewards from local files if database unavailable
        metrics["db_episode_count"] = len(p1_final_rewards)
        metrics["p1_final_mean"] = float(np.mean(p1_final_rewards))
        metrics["p1_final_std"] = float(np.std(p1_final_rewards))
        metrics["p1_final_se"] = calculate_standard_error(p1_final_rewards)
        metrics["p2_final_mean"] = float(np.mean(p2_final_rewards)) if p2_final_rewards else 0.0
        metrics["p2_final_std"] = float(np.std(p2_final_rewards)) if p2_final_rewards else 0.0
        metrics["p2_final_se"] = calculate_standard_error(p2_final_rewards) if p2_final_rewards else 0.0
        # Store raw rewards for CI calculation
        metrics["p1_final_rewards_raw"] = p1_final_rewards
        metrics["p2_final_rewards_raw"] = p2_final_rewards

    return metrics


def _calculate_single_metrics(experiment_name: str, data: dict[str, Any]) -> dict[str, Any]:
    """Calculate metrics for a single scenario experiment."""
    metrics: dict[str, Any] = {
        "tag": experiment_name,
        "experiment_type": "single",
        "scenario_id": data.get("scenario_id", "unknown"),
        "optimize_mode": data.get("optimize_mode", "unknown"),
        "bandit_type": data.get("bandit_type", "unknown"),
        "total_turns": data.get("total_turns", 0),
    }

    # Extract turn rewards
    turn_rewards = data.get("turn_rewards", [])
    if not turn_rewards:
        return {"tag": experiment_name, "error": "No turn rewards data"}

    p1_rewards = [t.get("p1_reward", 0) for t in turn_rewards if t.get("p1_reward") is not None]
    p2_rewards = [t.get("p2_reward", 0) for t in turn_rewards if t.get("p2_reward") is not None]

    metrics["valid_turns"] = len(p1_rewards)
    metrics["p1_overall_mean"] = float(np.mean(p1_rewards)) if p1_rewards else 0.0
    metrics["p1_overall_std"] = float(np.std(p1_rewards)) if p1_rewards else 0.0
    metrics["p1_overall_se"] = calculate_standard_error(p1_rewards)
    metrics["p2_overall_mean"] = float(np.mean(p2_rewards)) if p2_rewards else 0.0
    metrics["p2_overall_std"] = float(np.std(p2_rewards)) if p2_rewards else 0.0
    metrics["p2_overall_se"] = calculate_standard_error(p2_rewards)
    metrics["avg_overall"] = float(np.mean(p1_rewards + p2_rewards)) if (p1_rewards + p2_rewards) else 0.0
    metrics["avg_overall_se"] = calculate_standard_error(p1_rewards + p2_rewards)

    # Get final rewards if available
    metrics["p1_final_reward"] = p1_rewards[-1] if p1_rewards else 0.0
    metrics["p2_final_reward"] = p2_rewards[-1] if p2_rewards else 0.0

    # Arm selection info
    p1_bandit = data.get("p1_bandit")
    if p1_bandit and isinstance(p1_bandit, dict):
        metrics["p1_selections"] = p1_bandit.get("total_selections", 0)
        selections = p1_bandit.get("selections", [])
        if selections:
            arms_selected = [s.get("arm_index", 0) for s in selections]
            metrics["p1_unique_arms"] = len(set(arms_selected))

    p2_bandit = data.get("p2_bandit")
    if p2_bandit and isinstance(p2_bandit, dict):
        metrics["p2_selections"] = p2_bandit.get("total_selections", 0)
        selections = p2_bandit.get("selections", [])
        if selections:
            arms_selected = [s.get("arm_index", 0) for s in selections]
            metrics["p2_unique_arms"] = len(set(arms_selected))

    return metrics


def display_metrics(metrics: dict[str, Any]) -> None:
    """Display metrics in a formatted table."""
    if "error" in metrics:
        console.print(f"[red]Error for experiment '{metrics['tag']}': {metrics['error']}[/]")
        return

    is_batch = metrics.get("experiment_type") == "batch"

    # Summary panel
    if is_batch:
        summary_lines = [
            f"Experiment: [bold cyan]{metrics['tag']}[/]",
            f"Type: [bold yellow]Batch Experiment[/]",
            f"Subset: {metrics.get('subset', 'unknown')}",
            f"Bandit Type: [bold green]{metrics.get('bandit_type', 'unknown')}[/]",
            f"Optimize Mode: {metrics.get('optimize_mode', 'unknown')}",
            f"Episodes: {metrics.get('success_count', 0)}/{metrics.get('total_episodes', 0)}",
            f"Total Turns: {metrics.get('total_turns', 0)}",
            f"Avg Turns/Episode: {metrics.get('avg_turns_per_episode', 0):.1f}",
            f"Total Duration: {metrics.get('total_duration_seconds', 0)/60:.1f} min",
        ]
    else:
        summary_lines = [
            f"Experiment: [bold cyan]{metrics['tag']}[/]",
            f"Type: [bold yellow]Single Scenario[/]",
            f"Scenario: {metrics.get('scenario_id', 'unknown')}",
            f"Bandit Type: [bold green]{metrics.get('bandit_type', 'unknown')}[/]",
            f"Optimize Mode: {metrics.get('optimize_mode', 'unknown')}",
            f"Total Turns: {metrics.get('total_turns', 0)}",
        ]
    console.print(Panel("\n".join(summary_lines), title="Experiment Summary"))

    # Overall scores table - show per-turn rewards
    overall_table = Table(title="Per-Turn Reward Statistics")
    overall_table.add_column("Agent", style="cyan")
    overall_table.add_column("Mean", justify="right")
    overall_table.add_column("Std", justify="right", style="dim")
    overall_table.add_column("SE", justify="right", style="dim magenta")

    overall_table.add_row(
        "P1",
        f"{metrics['p1_overall_mean']:.4f}",
        f"{metrics['p1_overall_std']:.4f}",
        f"{metrics.get('p1_overall_se', 0):.4f}",
    )
    overall_table.add_row(
        "P2",
        f"{metrics['p2_overall_mean']:.4f}",
        f"{metrics['p2_overall_std']:.4f}",
        f"{metrics.get('p2_overall_se', 0):.4f}",
    )
    overall_table.add_row(
        "[bold]Average[/]",
        f"[bold]{metrics['avg_overall']:.4f}[/]",
        "",
        f"{metrics.get('avg_overall_se', 0):.4f}",
    )
    console.print(overall_table)

    # Final rewards from database (if available)
    if metrics.get("p1_final_mean") is not None or metrics.get("p1_final_reward") is not None:
        final_table = Table(title="Final Rewards (from Database)")
        final_table.add_column("Agent", style="cyan")
        final_table.add_column("Mean", justify="right", style="bold")
        final_table.add_column("Std", justify="right", style="dim")
        final_table.add_column("SE", justify="right", style="dim magenta")

        if is_batch and metrics.get("p1_final_mean") is not None:
            final_table.add_row(
                "P1",
                f"{metrics.get('p1_final_mean', 0):.4f}",
                f"{metrics.get('p1_final_std', 0):.4f}",
                f"{metrics.get('p1_final_se', 0):.4f}",
            )
            final_table.add_row(
                "P2",
                f"{metrics.get('p2_final_mean', 0):.4f}",
                f"{metrics.get('p2_final_std', 0):.4f}",
                f"{metrics.get('p2_final_se', 0):.4f}",
            )
            avg_final = (metrics.get('p1_final_mean', 0) + metrics.get('p2_final_mean', 0)) / 2
            # Calculate average SE (using pooled SE formula)
            p1_se = metrics.get('p1_final_se', 0)
            p2_se = metrics.get('p2_final_se', 0)
            avg_se = ((p1_se ** 2 + p2_se ** 2) ** 0.5) / 2  # Approximate pooled SE
            final_table.add_row(
                "[bold]Average[/]",
                f"[bold]{avg_final:.4f}[/]",
                "",
                f"{avg_se:.4f}",
            )
            console.print(final_table)
            console.print(f"[dim]Based on {metrics.get('db_episode_count', 0)} episodes from database[/]")
        elif not is_batch and metrics.get("p1_final_reward") is not None:
            final_table.add_row(
                "P1",
                f"{metrics.get('p1_final_reward', 0):.4f}",
                "-",
            )
            final_table.add_row(
                "P2",
                f"{metrics.get('p2_final_reward', 0):.4f}",
                "-",
            )
            console.print(final_table)

    # Show breakdown if available
    if metrics.get("p1_breakdown"):
        breakdown_table = Table(title="P1 Reward Breakdown (from Database)")
        breakdown_table.add_column("Dimension", style="cyan")
        breakdown_table.add_column("Score", justify="right")
        for dim, score in metrics["p1_breakdown"].items():
            breakdown_table.add_row(dim, f"{score:.4f}")
        console.print(breakdown_table)

    if metrics.get("p2_breakdown"):
        breakdown_table = Table(title="P2 Reward Breakdown (from Database)")
        breakdown_table.add_column("Dimension", style="cyan")
        breakdown_table.add_column("Score", justify="right")
        for dim, score in metrics["p2_breakdown"].items():
            breakdown_table.add_row(dim, f"{score:.4f}")
        console.print(breakdown_table)

    # Bandit selection info (only for single experiments)
    if not is_batch and (metrics.get("p1_selections") or metrics.get("p2_selections")):
        bandit_table = Table(title="Bandit Selection Info")
        bandit_table.add_column("Agent", style="cyan")
        bandit_table.add_column("Selections", justify="right")
        bandit_table.add_column("Unique Arms", justify="right")

        if metrics.get("p1_selections"):
            bandit_table.add_row(
                "P1",
                str(metrics.get("p1_selections", 0)),
                str(metrics.get("p1_unique_arms", 0))
            )
        if metrics.get("p2_selections"):
            bandit_table.add_row(
                "P2",
                str(metrics.get("p2_selections", 0)),
                str(metrics.get("p2_unique_arms", 0))
            )
        console.print(bandit_table)


def compare_experiments(
    experiments: list[str],
    outputs_dir: Path = OUTPUTS_DIR,
    use_ci: bool = False,
    max_workers: int = 8,
    env_ids_filter: set[str] | None = None
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Compare metrics across multiple experiments.

    Args:
        experiments: List of experiment names to compare
        outputs_dir: Directory containing experiment outputs
        use_ci: If True, calculate 95% confidence intervals instead of using std
        max_workers: Number of parallel workers for loading experiment data
        env_ids_filter: Optional set of environment IDs to filter episodes by.
                        If None, all episodes are included.

    Returns:
        Tuple of (main_df, dimensions_df) where dimensions_df has per-dimension breakdown
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    all_metrics = []
    dimension_rows = []

    # Parallel loading of experiment metrics
    def load_experiment(exp: str) -> tuple[str, dict[str, Any]]:
        return exp, calculate_metrics_for_experiment(exp, outputs_dir, env_ids_filter=env_ids_filter)

    console.print(f"[bold cyan]Loading {len(experiments)} experiments with {max_workers} workers...[/]")

    experiment_metrics: dict[str, dict[str, Any]] = {}
    loaded_count = 0
    error_count = 0
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(load_experiment, exp): exp for exp in experiments}
        for future in as_completed(futures):
            exp, metrics = future.result()
            experiment_metrics[exp] = metrics
            loaded_count += 1
            if "error" in metrics:
                error_count += 1
                console.print(f"  [red]✗[/] [{loaded_count}/{len(experiments)}] {exp}: {metrics.get('error', 'Unknown error')}")
            else:
                source = metrics.get("rewards_source", "local")
                source_tag = "[cyan](db)[/]" if source == "database" else "[dim](local)[/]"
                console.print(f"  [green]✓[/] [{loaded_count}/{len(experiments)}] {exp} {source_tag}")

    if error_count > 0:
        console.print(f"\n[yellow]⚠[/] Loaded {loaded_count - error_count}/{len(experiments)} experiments ({error_count} errors)\n")
    else:
        console.print(f"\n[green]✓[/] Loaded all {len(experiment_metrics)} experiments\n")

    # Process metrics in original order
    for exp in experiments:
        metrics = experiment_metrics.get(exp, {"error": "Not loaded"})
        if "error" not in metrics:
            is_batch = metrics.get("experiment_type") == "batch"
            row = {
                "experiment": exp,
                "type": "batch" if is_batch else "single",
                "bandit_type": metrics.get("bandit_type", "unknown"),
                "optimize": metrics.get("optimize_mode", "unknown"),
                "eta": metrics.get("eta", "N/A"),
                "episodes": metrics.get("success_count", 1) if is_batch else 1,
                "db_episodes": metrics.get("db_episode_count", 0),
                "turns": metrics.get("total_turns", 0),
                "p1_mean": metrics["p1_overall_mean"],
                "p1_std": metrics.get("p1_overall_std", 0),
                "p1_se": metrics.get("p1_overall_se", 0),
                "p2_mean": metrics["p2_overall_mean"],
                "p2_std": metrics.get("p2_overall_std", 0),
                "p2_se": metrics.get("p2_overall_se", 0),
                "avg": metrics["avg_overall"],
                "avg_se": metrics.get("avg_overall_se", 0),
                # Final rewards from database
                "p1_final": metrics.get("p1_final_mean", metrics.get("p1_final_reward", 0)),
                "p1_final_std": metrics.get("p1_final_std", 0),
                "p1_final_se": metrics.get("p1_final_se", 0),
                "p2_final": metrics.get("p2_final_mean", metrics.get("p2_final_reward", 0)),
                "p2_final_std": metrics.get("p2_final_std", 0),
                "p2_final_se": metrics.get("p2_final_se", 0),
                "avg_final": (metrics.get("p1_final_mean", 0) + metrics.get("p2_final_mean", 0)) / 2 if metrics.get("p1_final_mean") else 0,
            }
            
            # Calculate CI if requested and raw data is available
            if use_ci:
                # Use raw rewards from metrics (either from database or local files)
                p1_rewards_raw = metrics.get("p1_final_rewards_raw", [])
                p2_rewards_raw = metrics.get("p2_final_rewards_raw", [])
                if p1_rewards_raw:
                    p1_mean, p1_ci = calculate_confidence_interval(p1_rewards_raw)
                    p2_mean, p2_ci = calculate_confidence_interval(p2_rewards_raw)
                    row["p1_final_ci"] = p1_ci
                    row["p2_final_ci"] = p2_ci
                else:
                    # No CI available
                    row["p1_final_ci"] = 0.0
                    row["p2_final_ci"] = 0.0
            else:
                row["p1_final_ci"] = 0.0
                row["p2_final_ci"] = 0.0
            
            # Add goal scores (always included)
            row["p1_goal"] = metrics.get("p1_goal", 0.0)
            row["p2_goal"] = metrics.get("p2_goal", 0.0)

            # Add all dimension scores from breakdown
            p1_breakdown = metrics.get("p1_breakdown", {})
            p2_breakdown = metrics.get("p2_breakdown", {})
            # Add relationship and knowledge (and other dimensions)
            row["p1_relationship"] = p1_breakdown.get("relationship", 0.0)
            row["p2_relationship"] = p2_breakdown.get("relationship", 0.0)
            row["p1_knowledge"] = p1_breakdown.get("knowledge", 0.0)
            row["p2_knowledge"] = p2_breakdown.get("knowledge", 0.0)
            row["p1_believability"] = p1_breakdown.get("believability", 0.0)
            row["p2_believability"] = p2_breakdown.get("believability", 0.0)
            row["p1_secret"] = p1_breakdown.get("secret", 0.0)
            row["p2_secret"] = p2_breakdown.get("secret", 0.0)
            row["p1_social_rules"] = p1_breakdown.get("social_rules", 0.0)
            row["p2_social_rules"] = p2_breakdown.get("social_rules", 0.0)
            row["p1_financial_and_material_benefits"] = p1_breakdown.get("financial_and_material_benefits", 0.0)
            row["p2_financial_and_material_benefits"] = p2_breakdown.get("financial_and_material_benefits", 0.0)

            # Calculate goal SE and CI from raw breakdown data
            p1_breakdown_raw = metrics.get("p1_breakdown_raw", {})
            p2_breakdown_raw = metrics.get("p2_breakdown_raw", {})
            p1_breakdown_se = metrics.get("p1_breakdown_se", {})
            p2_breakdown_se = metrics.get("p2_breakdown_se", {})

            # Goal SE - always calculate if raw data available, otherwise use from metrics
            if p1_breakdown_raw and "goal" in p1_breakdown_raw:
                p1_goal_values = p1_breakdown_raw["goal"]
                row["p1_goal_se"] = calculate_standard_error(p1_goal_values) if p1_goal_values else 0.0
            else:
                row["p1_goal_se"] = p1_breakdown_se.get("goal", 0.0)

            if p2_breakdown_raw and "goal" in p2_breakdown_raw:
                p2_goal_values = p2_breakdown_raw["goal"]
                row["p2_goal_se"] = calculate_standard_error(p2_goal_values) if p2_goal_values else 0.0
            else:
                row["p2_goal_se"] = p2_breakdown_se.get("goal", 0.0)

            if use_ci:
                # Calculate goal CI from raw breakdown data
                if p1_breakdown_raw and "goal" in p1_breakdown_raw:
                    p1_goal_values = p1_breakdown_raw["goal"]
                    if p1_goal_values:
                        _, p1_goal_ci = calculate_confidence_interval(p1_goal_values)
                        row["p1_goal_ci"] = p1_goal_ci
                    else:
                        row["p1_goal_ci"] = 0.0
                else:
                    row["p1_goal_ci"] = 0.0

                if p2_breakdown_raw and "goal" in p2_breakdown_raw:
                    p2_goal_values = p2_breakdown_raw["goal"]
                    if p2_goal_values:
                        _, p2_goal_ci = calculate_confidence_interval(p2_goal_values)
                        row["p2_goal_ci"] = p2_goal_ci
                    else:
                        row["p2_goal_ci"] = 0.0
                else:
                    row["p2_goal_ci"] = 0.0
            
            # Add data quality information
            row["p1_float_count"] = metrics.get("p1_float_count", 0)
            row["p1_tuple_count"] = metrics.get("p1_tuple_count", 0)
            row["p2_float_count"] = metrics.get("p2_float_count", 0)
            row["p2_tuple_count"] = metrics.get("p2_tuple_count", 0)
            row["data_quality_status"] = metrics.get("data_quality_status", "unknown")
            
            all_metrics.append(row)

            # Collect dimension breakdowns
            p1_breakdown = metrics.get("p1_breakdown", {})
            p1_breakdown_std = metrics.get("p1_breakdown_std", {})
            p1_breakdown_se = metrics.get("p1_breakdown_se", {})
            p2_breakdown = metrics.get("p2_breakdown", {})
            p2_breakdown_std = metrics.get("p2_breakdown_std", {})
            p2_breakdown_se = metrics.get("p2_breakdown_se", {})
            p1_breakdown_raw = metrics.get("p1_breakdown_raw", {})
            p2_breakdown_raw = metrics.get("p2_breakdown_raw", {})

            if p1_breakdown or p2_breakdown:
                dim_row = {
                    "experiment": exp,
                    "bandit_type": metrics.get("bandit_type", "unknown"),
                    # Add data quality info for coloring
                    "p1_float_count": metrics.get("p1_float_count", 0),
                    "p2_float_count": metrics.get("p2_float_count", 0),
                    "p1_tuple_count": metrics.get("p1_tuple_count", 0),
                    "p2_tuple_count": metrics.get("p2_tuple_count", 0),
                    "db_episodes": metrics.get("db_episode_count", 70),
                    "data_quality_status": metrics.get("data_quality_status", "unknown"),
                }
                # Add each dimension for P1 and P2
                for dim in DIMENSIONS:
                    dim_row[f"p1_{dim}"] = p1_breakdown.get(dim, 0)
                    dim_row[f"p2_{dim}"] = p2_breakdown.get(dim, 0)
                    # Always include standard error and standard deviation
                    dim_row[f"p1_{dim}_se"] = p1_breakdown_se.get(dim, 0)
                    dim_row[f"p2_{dim}_se"] = p2_breakdown_se.get(dim, 0)
                    dim_row[f"p1_{dim}_std"] = p1_breakdown_std.get(dim, 0)
                    dim_row[f"p2_{dim}_std"] = p2_breakdown_std.get(dim, 0)

                    if use_ci and p1_breakdown_raw and p2_breakdown_raw:
                        # Calculate CI for this dimension
                        p1_dim_values = p1_breakdown_raw.get(dim, [])
                        p2_dim_values = p2_breakdown_raw.get(dim, [])

                        if p1_dim_values:
                            _, p1_ci = calculate_confidence_interval(p1_dim_values)
                            dim_row[f"p1_{dim}_ci"] = p1_ci
                        else:
                            dim_row[f"p1_{dim}_ci"] = 0.0

                        if p2_dim_values:
                            _, p2_ci = calculate_confidence_interval(p2_dim_values)
                            dim_row[f"p2_{dim}_ci"] = p2_ci
                        else:
                            dim_row[f"p2_{dim}_ci"] = 0.0

                # Overall score
                dim_row["p1_overall"] = p1_breakdown.get("overall_score", metrics.get("p1_final_mean", 0))
                dim_row["p2_overall"] = p2_breakdown.get("overall_score", metrics.get("p2_final_mean", 0))
                # Always include SE and Std for overall
                dim_row["p1_overall_se"] = p1_breakdown_se.get("overall_score", metrics.get("p1_final_se", 0))
                dim_row["p2_overall_se"] = p2_breakdown_se.get("overall_score", metrics.get("p2_final_se", 0))
                dim_row["p1_overall_std"] = p1_breakdown_std.get("overall_score", metrics.get("p1_final_std", 0))
                dim_row["p2_overall_std"] = p2_breakdown_std.get("overall_score", metrics.get("p2_final_std", 0))

                if use_ci and p1_breakdown_raw and p2_breakdown_raw:
                    # Calculate CI for overall score
                    p1_overall_values = p1_breakdown_raw.get("overall_score", [])
                    p2_overall_values = p2_breakdown_raw.get("overall_score", [])

                    if p1_overall_values:
                        _, p1_overall_ci = calculate_confidence_interval(p1_overall_values)
                        dim_row["p1_overall_ci"] = p1_overall_ci
                    else:
                        dim_row["p1_overall_ci"] = 0.0

                    if p2_overall_values:
                        _, p2_overall_ci = calculate_confidence_interval(p2_overall_values)
                        dim_row["p2_overall_ci"] = p2_overall_ci
                    else:
                        dim_row["p2_overall_ci"] = 0.0

                dimension_rows.append(dim_row)
        else:
            logger.warning(f"Skipping {exp}: {metrics.get('error')}")

    if not all_metrics:
        logger.error("No valid metrics to compare")
        return pd.DataFrame(), pd.DataFrame()

    df = pd.DataFrame(all_metrics)
    dim_df = pd.DataFrame(dimension_rows) if dimension_rows else pd.DataFrame()
    return df, dim_df


def _get_multi_evalset_metric_value_and_se(
    row: pd.Series,
    dim_row: pd.Series | dict[str, Any],
    evalset: str,
    metric: str,
) -> tuple[float, float]:
    """Get the display value and SE for a metric in multi-eval-set comparison."""
    if metric == "goal":
        p1_value = row.get("p1_goal", row.get("p1_final", 0)) or 0
        p2_value = row.get("p2_goal", row.get("p2_final", 0)) or 0
        value = (p1_value + p2_value) / 2
        if evalset == "hard":
            p1_std = dim_row.get("p1_goal_std", row.get("p1_final_std", 0)) or 0
            p2_std = dim_row.get("p2_goal_std", row.get("p2_final_std", 0)) or 0
            n = dim_row.get("db_episodes", row.get("db_episode_count", 1)) or 1
            se = ((p1_std ** 2 + p2_std ** 2) ** 0.5) / 2 / n if n > 0 else 0
        else:
            p1_se = dim_row.get("p1_goal_se", row.get("p1_final_se", 0)) or 0
            p2_se = dim_row.get("p2_goal_se", row.get("p2_final_se", 0)) or 0
            se = ((p1_se ** 2 + p2_se ** 2) ** 0.5) / 2
        return value, se

    if metric == "relationship":
        p1_value = row.get("p1_relationship", 0) or 0
        p2_value = row.get("p2_relationship", 0) or 0
        value = (p1_value + p2_value) / 2
        if evalset == "hard":
            p1_std = dim_row.get("p1_relationship_std", 0) or 0
            p2_std = dim_row.get("p2_relationship_std", 0) or 0
            n = dim_row.get("db_episodes", row.get("db_episode_count", 1)) or 1
            se = ((p1_std ** 2 + p2_std ** 2) ** 0.5) / 2 / n if n > 0 else 0
        else:
            p1_se = dim_row.get("p1_relationship_se", 0) or 0
            p2_se = dim_row.get("p2_relationship_se", 0) or 0
            se = ((p1_se ** 2 + p2_se ** 2) ** 0.5) / 2
        return value, se

    if metric == "knowledge":
        p1_value = row.get("p1_knowledge", 0) or 0
        p2_value = row.get("p2_knowledge", 0) or 0
        value = (p1_value + p2_value) / 2
        if evalset == "hard":
            p1_std = dim_row.get("p1_knowledge_std", 0) or 0
            p2_std = dim_row.get("p2_knowledge_std", 0) or 0
            n = dim_row.get("db_episodes", row.get("db_episode_count", 1)) or 1
            se = ((p1_std ** 2 + p2_std ** 2) ** 0.5) / 2 / n if n > 0 else 0
        else:
            p1_se = dim_row.get("p1_knowledge_se", 0) or 0
            p2_se = dim_row.get("p2_knowledge_se", 0) or 0
            se = ((p1_se ** 2 + p2_se ** 2) ** 0.5) / 2
        return value, se

    if metric == "overall":
        value = row.get("avg_final", row.get("avg", 0)) or 0
        if evalset == "hard":
            p1_std = dim_row.get("p1_overall_std", row.get("p1_final_std", 0)) or 0
            p2_std = dim_row.get("p2_overall_std", row.get("p2_final_std", 0)) or 0
            n = dim_row.get("db_episodes", row.get("db_episode_count", 1)) or 1
            se = ((p1_std ** 2 + p2_std ** 2) ** 0.5) / 2 / n if n > 0 else 0
        else:
            p1_se = dim_row.get("p1_overall_se", row.get("p1_final_se", 0)) or 0
            p2_se = dim_row.get("p2_overall_se", row.get("p2_final_se", 0)) or 0
            se = ((p1_se ** 2 + p2_se ** 2) ** 0.5) / 2
        return value, se

    raise ValueError(f"Unsupported metric: {metric}")


def build_summary_dataframe(
    results_by_evalset: dict[str, tuple[pd.DataFrame, pd.DataFrame]],
) -> pd.DataFrame:
    """Build the copy-friendly summary as a DataFrame."""
    if not results_by_evalset:
        return pd.DataFrame()

    first_key = list(results_by_evalset.keys())[0]
    first_df = results_by_evalset[first_key][0]
    if first_df.empty:
        return pd.DataFrame()

    experiments = first_df["experiment"].tolist()
    use_evalset_prefix = len(results_by_evalset) > 1
    summary_rows: list[dict[str, Any]] = []

    for exp in experiments:
        summary_row: dict[str, Any] = {}
        for evalset, (df, dim_df) in results_by_evalset.items():
            display_name = EVALSET_DISPLAY_NAMES.get(evalset, evalset.upper())
            exp_row = df[df["experiment"] == exp]
            dim_exp_row = dim_df[dim_df["experiment"] == exp] if not dim_df.empty else pd.DataFrame()
            if exp_row.empty:
                for _, metric_label in SUMMARY_METRICS:
                    value_column = f"{display_name} {metric_label}" if use_evalset_prefix else metric_label
                    se_column = f"{display_name} {metric_label} SE" if use_evalset_prefix else f"{metric_label} SE"
                    summary_row[value_column] = "N/A"
                    summary_row[se_column] = "N/A"
                continue

            row = exp_row.iloc[0]
            dim_row = dim_exp_row.iloc[0] if not dim_exp_row.empty else {}
            for metric_key, metric_label in SUMMARY_METRICS:
                value, se = _get_multi_evalset_metric_value_and_se(row, dim_row, evalset, metric_key)
                value_column = f"{display_name} {metric_label}" if use_evalset_prefix else metric_label
                se_column = f"{display_name} {metric_label} SE" if use_evalset_prefix else f"{metric_label} SE"
                summary_row[value_column] = round(value, 4)
                summary_row[se_column] = round(se, 4)

        summary_row["tag"] = exp
        summary_rows.append(summary_row)

    for index, summary_row in enumerate(summary_rows, start=1):
        summary_row["No."] = index

    ordered_columns: list[str] = ["No."]
    for evalset in results_by_evalset.keys():
        display_name = EVALSET_DISPLAY_NAMES.get(evalset, evalset.upper())
        for _, metric_label in SUMMARY_METRICS:
            ordered_columns.append(f"{display_name} {metric_label}" if use_evalset_prefix else metric_label)
            ordered_columns.append(f"{display_name} {metric_label} SE" if use_evalset_prefix else f"{metric_label} SE")
    ordered_columns.append("tag")

    return pd.DataFrame(summary_rows, columns=ordered_columns)


def export_multi_evalset_summary_xlsx(summary_df: pd.DataFrame, output_path: str | Path) -> None:
    """Export the multi-eval-set summary DataFrame to XLSX."""
    try:
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "XLSX export requires the 'openpyxl' package. Install dependencies and retry."
        ) from exc

    highlight_fills = [
        PatternFill(fill_type="solid", fgColor="FFF2CC"),
        PatternFill(fill_type="solid", fgColor="D9EAD3"),
        PatternFill(fill_type="solid", fgColor="D0E0E3"),
    ]

    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            summary_df.to_excel(writer, index=False, sheet_name="summary")
            worksheet = writer.book["summary"]

            for column_index, column_name in enumerate(summary_df.columns, start=1):
                if column_name in {"No.", "tag"} or column_name.endswith(" SE"):
                    continue

                numeric_series = pd.to_numeric(summary_df[column_name], errors="coerce")
                ranked_indices = numeric_series.sort_values(ascending=False).dropna().head(3).index.tolist()
                for rank_index, row_index in enumerate(ranked_indices):
                    cell = worksheet[f"{get_column_letter(column_index)}{row_index + 2}"]
                    cell.fill = highlight_fills[rank_index]
    except ImportError as exc:
        raise RuntimeError(
            "XLSX export requires the 'openpyxl' package. Install dependencies and retry."
        ) from exc


def display_multi_evalset_comparison(
    results_by_evalset: dict[str, tuple[pd.DataFrame, pd.DataFrame]],
    use_ci: bool = False
) -> None:
    """Display a side-by-side comparison table for multiple eval-sets.

    Generates a table like:
    SOTOPIA                 SOTOPIA-Hard
    Goal ↑   Overall ↑     Goal ↑   Overall ↑     Experiment
    0.85     0.72          0.65     0.58          exp1
    0.88     0.75          0.70     0.62          exp2

    Args:
        results_by_evalset: Dict mapping eval-set name to (df, dim_df) tuple
        use_ci: If True, display confidence intervals
    """
    if not results_by_evalset:
        console.print("[red]No data to compare[/]")
        return

    # Get experiment list from the first eval-set
    first_key = list(results_by_evalset.keys())[0]
    first_df = results_by_evalset[first_key][0]

    if first_df.empty:
        console.print("[red]No data to compare[/]")
        return

    experiments = first_df["experiment"].tolist()

    # Build the table
    console.print("\n[bold cyan]═══ Multi-Dataset Comparison ═══[/]\n")

    # Create table with dynamic columns based on eval-sets
    table = Table(title="Goal & Overall Score by Dataset", show_header=True, header_style="bold")

    # Add columns for each eval-set FIRST
    for evalset in results_by_evalset.keys():
        display_name = EVALSET_DISPLAY_NAMES.get(evalset, evalset.upper())
        table.add_column(f"{display_name}\nGoal ↑", justify="right", style="green")
        table.add_column(f"{display_name}\nGoal SE", justify="right", style="dim green")
        table.add_column(f"{display_name}\nRelation ↑", justify="right", style="blue")
        table.add_column(f"{display_name}\nRelation SE", justify="right", style="dim blue")
        table.add_column(f"{display_name}\nKnowledge ↑", justify="right", style="magenta")
        table.add_column(f"{display_name}\nKnowledge SE", justify="right", style="dim magenta")
        table.add_column(f"{display_name}\nOverall ↑", justify="right", style="yellow")
        table.add_column(f"{display_name}\nOverall SE", justify="right", style="dim yellow")

    # Add Experiment column LAST
    table.add_column("Experiment", style="cyan", no_wrap=False)

    # First pass: collect all metric values for ranking
    # Structure: {(evalset, metric_name): {exp: value}}
    metric_values: dict[tuple, dict[str, float]] = {}
    for evalset in results_by_evalset.keys():
        for metric_name in ["goal", "relationship", "knowledge", "overall"]:
            metric_values[(evalset, metric_name)] = {}

    for exp in experiments:
        for evalset, (df, dim_df) in results_by_evalset.items():
            exp_row = df[df["experiment"] == exp]
            if not exp_row.empty:
                row = exp_row.iloc[0]
                goal, _ = _get_multi_evalset_metric_value_and_se(row, {}, evalset, "goal")
                relationship, _ = _get_multi_evalset_metric_value_and_se(row, {}, evalset, "relationship")
                knowledge, _ = _get_multi_evalset_metric_value_and_se(row, {}, evalset, "knowledge")
                overall, _ = _get_multi_evalset_metric_value_and_se(row, {}, evalset, "overall")
                metric_values[(evalset, "goal")][exp] = goal
                metric_values[(evalset, "relationship")][exp] = relationship
                metric_values[(evalset, "knowledge")][exp] = knowledge
                metric_values[(evalset, "overall")][exp] = overall

    # Compute rankings for each column (higher is better)
    # Structure: {(evalset, metric_name): {exp: rank}} where rank is 1, 2, 3 or None
    rankings: dict[tuple, dict[str, int | None]] = {}
    for key, exp_values in metric_values.items():
        sorted_exps = sorted(exp_values.items(), key=lambda x: x[1], reverse=True)
        rankings[key] = {}
        for rank, (exp, _) in enumerate(sorted_exps[:3], start=1):
            rankings[key][exp] = rank

    def format_with_rank(value: float, exp: str, evalset: str, metric: str) -> str:
        """Format value with ranking indicator if in top 3."""
        rank = rankings.get((evalset, metric), {}).get(exp)
        if rank:
            return f"{value:.4f} [{rank}]"
        return f"{value:.4f}"

    # Add rows for each experiment
    for exp in experiments:
        row_data: list[str] = []

        for evalset, (df, dim_df) in results_by_evalset.items():
            exp_row = df[df["experiment"] == exp]
            dim_exp_row = dim_df[dim_df["experiment"] == exp] if not dim_df.empty else pd.DataFrame()
            if not exp_row.empty:
                row = exp_row.iloc[0]
                dim_row = dim_exp_row.iloc[0] if not dim_exp_row.empty else {}
                goal, goal_se = _get_multi_evalset_metric_value_and_se(row, dim_row, evalset, "goal")
                relationship, relationship_se = _get_multi_evalset_metric_value_and_se(row, dim_row, evalset, "relationship")
                knowledge, knowledge_se = _get_multi_evalset_metric_value_and_se(row, dim_row, evalset, "knowledge")
                overall, overall_se = _get_multi_evalset_metric_value_and_se(row, dim_row, evalset, "overall")

                row_data.extend([
                    format_with_rank(goal, exp, evalset, "goal"),
                    f"{goal_se:.4f}",
                    format_with_rank(relationship, exp, evalset, "relationship"),
                    f"{relationship_se:.4f}",
                    format_with_rank(knowledge, exp, evalset, "knowledge"),
                    f"{knowledge_se:.4f}",
                    format_with_rank(overall, exp, evalset, "overall"),
                    f"{overall_se:.4f}",
                ])
            else:
                row_data.extend(["N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A"])

        # Add experiment name at the end (full name)
        row_data.append(exp)
        table.add_row(*row_data)

    console.print(table)

    # Also display a simplified summary table (just Goal and Overall)
    console.print("\n[bold cyan]═══ Summary Table (Copy-Friendly) ═══[/]\n")

    summary_df = build_summary_dataframe(results_by_evalset)
    summary_table = Table(show_header=True, header_style="bold", box=None)

    for column_name in summary_df.columns[:-1]:
        summary_table.add_column(column_name, justify="right")
    summary_table.add_column(summary_df.columns[-1], style="cyan")

    for _, summary_row in summary_df.iterrows():
        row_data = [
            value if isinstance(value, str) else f"{value:.4f}"
            for value in summary_row.tolist()
        ]
        summary_table.add_row(*row_data)

    console.print(summary_table)

    # Display compact all-dimensions table (similar to Goal & Overall table)
    # console.print("\n[bold cyan]═══ All Dimensions by Dataset (Avg of P1 & P2) ═══[/]\n")
    #
    # all_dims_by_dataset_table = Table(show_header=True, header_style="bold", box=None)
    #
    # # Add dimension columns for each eval-set
    # for evalset in results_by_evalset.keys():
    #     display_name = evalset_display_names.get(evalset, evalset.upper())
    #     for dim in DIMENSIONS:
    #         dim_name = dim.replace("_", " ").title()
    #         # Shorten long dimension names
    #         if dim_name == "Financial And Material Benefits":
    #             dim_name = "Financial"
    #         elif dim_name == "Social Rules":
    #             dim_name = "Social"
    #         all_dims_by_dataset_table.add_column(f"{display_name}\n{dim_name}", justify="right")
    #
    # # Add Experiment column last
    # all_dims_by_dataset_table.add_column("Experiment", style="cyan")
    #
    # for exp in experiments:
    #     row_data: list[str] = []
    #     for evalset, (_, dim_df) in results_by_evalset.items():
    #         if dim_df is not None and not dim_df.empty:
    #             exp_row = dim_df[dim_df["experiment"] == exp]
    #             if not exp_row.empty:
    #                 row = exp_row.iloc[0]
    #                 for dim in DIMENSIONS:
    #                     p1_val = row.get(f"p1_{dim}", 0) or 0
    #                     p2_val = row.get(f"p2_{dim}", 0) or 0
    #                     avg_val = (p1_val + p2_val) / 2
    #                     row_data.append(f"{avg_val:.4f}")
    #             else:
    #                 row_data.extend(["N/A"] * len(DIMENSIONS))
    #         else:
    #             row_data.extend(["N/A"] * len(DIMENSIONS))
    #     row_data.append(exp)
    #     all_dims_by_dataset_table.add_row(*row_data)
    #
    # console.print(all_dims_by_dataset_table)

    # Display all dimensions table (for multi-dataset comparison)
    # Shows average of P1 and P2, sorted by score descending
    # console.print("\n[bold cyan]═══ All Dimensions Comparison (Avg of P1 & P2, Sorted Descending) ═══[/]\n")
    #
    # # Create a table for each dimension showing average scores across datasets
    # for dim in DIMENSIONS:
    #     dim_table = Table(title=f"{dim.replace('_', ' ').title()} ↓", show_header=True, header_style="bold")
    #
    #     # Add columns: for each evalset, add Avg and SE columns
    #     for evalset in results_by_evalset.keys():
    #         display_name = evalset_display_names.get(evalset, evalset.upper())
    #         dim_table.add_column(f"{display_name}\nAvg", justify="right", style="green")
    #         dim_table.add_column(f"{display_name}\nSE", justify="right", style="dim")
    #
    #     # Add Experiment column last
    #     dim_table.add_column("Experiment", style="cyan", no_wrap=False)
    #
    #     # Collect data for sorting: (avg_score, exp, row_data)
    #     exp_scores: list[tuple[float, str, list[str]]] = []
    #     for exp in experiments:
    #         row_data: list[str] = []
    #         total_score = 0.0
    #         score_count = 0
    #         for evalset, (_, dim_df) in results_by_evalset.items():
    #             if dim_df is not None and not dim_df.empty:
    #                 exp_row = dim_df[dim_df["experiment"] == exp]
    #                 if not exp_row.empty:
    #                     row = exp_row.iloc[0]
    #                     p1_val = row.get(f"p1_{dim}", 0) or 0
    #                     p1_se = row.get(f"p1_{dim}_se", 0) or 0
    #                     p2_val = row.get(f"p2_{dim}", 0) or 0
    #                     p2_se = row.get(f"p2_{dim}_se", 0) or 0
    #                     # Calculate average of P1 and P2
    #                     avg_val = (p1_val + p2_val) / 2
    #                     # SE of average: sqrt((se1^2 + se2^2) / 4)
    #                     avg_se = math.sqrt((p1_se ** 2 + p2_se ** 2) / 4) if (p1_se or p2_se) else 0
    #                     row_data.extend([
    #                         f"{avg_val:.4f}",
    #                         f"{avg_se:.4f}",
    #                     ])
    #                     total_score += avg_val
    #                     score_count += 1
    #                 else:
    #                     row_data.extend(["N/A", "N/A"])
    #             else:
    #                 row_data.extend(["N/A", "N/A"])
    #         row_data.append(exp)
    #         avg_score = total_score / score_count if score_count > 0 else 0.0
    #         exp_scores.append((avg_score, exp, row_data))
    #
    #     # Sort by score descending
    #     exp_scores.sort(key=lambda x: x[0], reverse=True)
    #
    #     # Add sorted rows to table
    #     for _, _, row_data in exp_scores:
    #         dim_table.add_row(*row_data)
    #
    #     console.print(dim_table)
    #     console.print()  # Add spacing between dimension tables

    # Also display a compact all-dimensions summary table
    # console.print("\n[bold cyan]═══ All Dimensions Summary (Copy-Friendly) ═══[/]\n")
    #
    # all_dims_table = Table(show_header=True, header_style="bold", box=None)
    #
    # # Add dimension column first
    # all_dims_table.add_column("Dimension", style="magenta")
    #
    # # Add columns for each evalset (Avg only)
    # for evalset in results_by_evalset.keys():
    #     display_name = evalset_display_names.get(evalset, evalset.upper())
    #     all_dims_table.add_column(f"{display_name}", justify="right")
    #
    # # Add Experiment column last
    # all_dims_table.add_column("Experiment", style="cyan")
    #
    # # For each experiment, add rows for each dimension
    # for exp in experiments:
    #     for dim in DIMENSIONS:
    #         row_data: list[str] = [dim.replace("_", " ").title()]
    #         for evalset, (_, dim_df) in results_by_evalset.items():
    #             if dim_df is not None and not dim_df.empty:
    #                 exp_row = dim_df[dim_df["experiment"] == exp]
    #                 if not exp_row.empty:
    #                     row = exp_row.iloc[0]
    #                     p1_val = row.get(f"p1_{dim}", 0) or 0
    #                     p2_val = row.get(f"p2_{dim}", 0) or 0
    #                     avg_val = (p1_val + p2_val) / 2
    #                     row_data.append(f"{avg_val:.4f}")
    #                 else:
    #                     row_data.append("N/A")
    #             else:
    #                 row_data.append("N/A")
    #         # Only add experiment name on first dimension row
    #         if dim == DIMENSIONS[0]:
    #             row_data.append(exp)
    #         else:
    #             row_data.append("")
    #         all_dims_table.add_row(*row_data)
    #     # Add a separator row between experiments (empty row)
    #     if exp != experiments[-1]:
    #         all_dims_table.add_row(*["─" * 10 for _ in range(1 + len(results_by_evalset))] + ["─" * 20])
    #
    # console.print(all_dims_table)


def display_comparison(df: pd.DataFrame, dim_df: pd.DataFrame | None = None, use_ci: bool = False, export_dir: str | None = None) -> None:
    """Display comparison table.
    
    Args:
        df: DataFrame with experiment metrics
        dim_df: DataFrame with dimension breakdowns
        use_ci: If True, display confidence intervals instead of std for final rewards
        export_dir: If specified, export all tables to CSV files in this directory
    """
    if df.empty:
        console.print("[red]No data to compare[/]")
        return

    # Check if we have batch experiments and final rewards
    has_batch = "type" in df.columns and (df["type"] == "batch").any()
    
    # Ensure numeric types
    numeric_cols = ["p1_final", "p2_final", "avg_final", "p1_goal", "p2_goal"]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
            
    has_final = "p1_final" in df.columns # and df["p1_final"].sum() > 0  <-- Relaxed check
    has_ci = use_ci and "p1_final_ci" in df.columns
    
    # Create export directory if needed
    if export_dir:
        export_path = Path(export_dir)
        export_path.mkdir(parents=True, exist_ok=True)
        console.print(f"\n[green]Exporting tables to: {export_path}[/]\n")

    # Overall comparison - Per-Turn Rewards
    console.print("\n[bold cyan]═══ Experiment Comparison ═══[/]\n")

    table = Table(title="Per-Turn Reward Statistics")
    table.add_column("Experiment", style="cyan", no_wrap=True)
    table.add_column("Bandit", justify="center", style="green")
    table.add_column("Opt", justify="center")
    if has_batch:
        table.add_column("Episodes", justify="right")
    table.add_column("Turns", justify="right")
    table.add_column("P1 Mean", justify="right")
    table.add_column("P1 Std", justify="right", style="dim")
    table.add_column("P1 SE", justify="right", style="dim magenta")
    table.add_column("P2 Mean", justify="right")
    table.add_column("P2 SE", justify="right", style="dim magenta")
    table.add_column("Avg", justify="right", style="bold")
    table.add_column("Avg SE", justify="right", style="dim magenta")

    for _, row in df.iterrows():
        row_data = [
            str(row["experiment"]),
            str(row["bandit_type"]),
            str(row["optimize"]),
        ]
        if has_batch:
            row_data.append(str(int(row.get("episodes", 1))))
        row_data.extend([
            str(int(row["turns"])),
            f"{row['p1_mean']:.4f}",
            f"{row.get('p1_std', 0):.4f}",
            f"{row.get('p1_se', 0):.4f}",
            f"{row['p2_mean']:.4f}",
            f"{row.get('p2_se', 0):.4f}",
            f"{row['avg']:.4f}",
            f"{row.get('avg_se', 0):.4f}",
        ])
        table.add_row(*row_data)
    console.print(table)

    # Export Per-Turn Rewards to CSV
    if export_dir:
        csv_path = export_path / "per_turn_rewards.csv"
        export_cols = ["experiment", "bandit_type", "optimize"]
        if has_batch:
            export_cols.append("episodes")
        export_cols.extend(["turns", "p1_mean", "p1_std", "p1_se", "p2_mean", "p2_std", "p2_se", "avg", "avg_se"])
        available_cols = [c for c in export_cols if c in df.columns]
        df[available_cols].to_csv(csv_path, index=False)
        console.print(f"[dim]✓ Exported: {csv_path.name}[/]")

    # Final Rewards from Database
    if has_final:
        console.print("\n[bold cyan]═══ Final Rewards (from Database) ═══[/]\n")

        final_table = Table(title="Final Episode Rewards" + (" with 95% CI" if has_ci else ""))
        final_table.add_column("Experiment", style="cyan", no_wrap=True)
        final_table.add_column("Bandit", justify="center", style="green")
        final_table.add_column("Opt", justify="center")
        final_table.add_column("Eta", justify="center", style="magenta")
        final_table.add_column("DB Episodes", justify="right")
        
        # Add data quality column
        final_table.add_column("Data Quality", justify="center", style="yellow")
        
        if has_ci:
            # Show with CI: mean and CI in separate columns
            final_table.add_column("P1 Mean", justify="right", style="yellow")
            final_table.add_column("P1 CI (±)", justify="right", style="dim")
            final_table.add_column("P1 SE", justify="right", style="dim magenta")
            final_table.add_column("P2 Mean", justify="right", style="yellow")
            final_table.add_column("P2 CI (±)", justify="right", style="dim")
            final_table.add_column("P2 SE", justify="right", style="dim magenta")
            # Add goal score columns with CI and SE
            final_table.add_column("P1 Goal", justify="right", style="cyan")
            final_table.add_column("P1 Goal CI", justify="right", style="dim cyan")
            final_table.add_column("P1 Goal SE", justify="right", style="dim magenta")
            final_table.add_column("P2 Goal", justify="right", style="cyan")
            final_table.add_column("P2 Goal CI", justify="right", style="dim cyan")
            final_table.add_column("P2 Goal SE", justify="right", style="dim magenta")
        else:
            # Show with std and SE: mean, std, SE
            final_table.add_column("P1 Final", justify="right", style="yellow")
            final_table.add_column("P1 Std", justify="right", style="dim")
            final_table.add_column("P1 SE", justify="right", style="dim magenta")
            final_table.add_column("P2 Final", justify="right", style="yellow")
            final_table.add_column("P2 SE", justify="right", style="dim magenta")
            # Add goal score columns with SE
            final_table.add_column("P1 Goal", justify="right", style="cyan")
            final_table.add_column("P1 Goal SE", justify="right", style="dim magenta")
            final_table.add_column("P2 Goal", justify="right", style="cyan")
            final_table.add_column("P2 Goal SE", justify="right", style="dim magenta")
        final_table.add_column("Avg Final", justify="right", style="bold")
        final_table.add_column("Avg Goal", justify="right", style="bold cyan")

        for _, row in df.iterrows():
            # Get data quality info
            p1_float = row.get("p1_float_count", 0)
            p2_float = row.get("p2_float_count", 0)
            total_eps = row.get("db_episodes", 0)
            
            # Determine quality status
            if p1_float > 0 or p2_float > 0:
                float_pct = ((p1_float + p2_float) / (2 * total_eps) * 100) if total_eps > 0 else 0
                if float_pct >= 20:
                    quality_str = f"[bold red]✗ {p1_float + p2_float} float[/]"
                elif float_pct >= 5:
                    quality_str = f"[orange]⚠ {p1_float + p2_float} float[/]"
                else:
                    quality_str = f"[yellow]⚠ {p1_float + p2_float} float[/]"
            else:
                quality_str = "[green]✓[/]"
            
            # Color the experiment name based on data quality
            exp_name = str(row["experiment"])
            if p1_float > 0 or p2_float > 0:
                float_pct = ((p1_float + p2_float) / (2 * total_eps) * 100) if total_eps > 0 else 0
                if float_pct >= 20:
                    exp_name = f"[bold red]{exp_name}[/]"
                elif float_pct >= 5:
                    exp_name = f"[orange]{exp_name}[/]"
                else:
                    exp_name = f"[yellow]{exp_name}[/]"
            
            row_data = [
                exp_name,  # Colored experiment name
                str(row["bandit_type"]),
                str(row["optimize"]),
                str(row.get("eta", "N/A")),
                str(int(row.get("db_episodes", 0))),
                quality_str,  # Data quality status
            ]
            
            if has_ci:
                # Show mean, CI, and SE separately
                p1_ci = row.get("p1_final_ci", 0)
                p2_ci = row.get("p2_final_ci", 0)
                p1_se = row.get("p1_final_se", 0)
                p2_se = row.get("p2_final_se", 0)
                p1_goal = row.get("p1_goal", 0)
                p2_goal = row.get("p2_goal", 0)
                p1_goal_ci = row.get("p1_goal_ci", 0)
                p2_goal_ci = row.get("p2_goal_ci", 0)
                p1_goal_se = row.get("p1_goal_se", 0)
                p2_goal_se = row.get("p2_goal_se", 0)
                row_data.extend([
                    f"{row.get('p1_final', 0):.4f}",
                    f"±{p1_ci:.4f}",
                    f"{p1_se:.4f}",
                    f"{row.get('p2_final', 0):.4f}",
                    f"±{p2_ci:.4f}",
                    f"{p2_se:.4f}",
                    f"{p1_goal:.2f}",
                    f"±{p1_goal_ci:.2f}",
                    f"{p1_goal_se:.4f}",
                    f"{p2_goal:.2f}",
                    f"±{p2_goal_ci:.2f}",
                    f"{p2_goal_se:.4f}",
                ])
            else:
                # Format with std and SE, including goal
                p1_goal = row.get("p1_goal", 0)
                p2_goal = row.get("p2_goal", 0)
                p1_goal_se = row.get("p1_goal_se", 0)
                p2_goal_se = row.get("p2_goal_se", 0)
                row_data.extend([
                    f"{row.get('p1_final', 0):.4f}",
                    f"{row.get('p1_final_std', 0):.4f}",
                    f"{row.get('p1_final_se', 0):.4f}",
                    f"{row.get('p2_final', 0):.4f}",
                    f"{row.get('p2_final_se', 0):.4f}",
                    f"{p1_goal:.2f}",
                    f"{p1_goal_se:.4f}",
                    f"{p2_goal:.2f}",
                    f"{p2_goal_se:.4f}",
                ])

            row_data.append(f"{row.get('avg_final', 0):.4f}")
            # Calculate avg_goal for this row
            p1_goal_val = row.get("p1_goal", 0) or 0
            p2_goal_val = row.get("p2_goal", 0) or 0
            avg_goal_val = (p1_goal_val + p2_goal_val) / 2
            row_data.append(f"{avg_goal_val:.2f}")
            final_table.add_row(*row_data)
        console.print(final_table)

        # Export Final Rewards to CSV
        if export_dir:
            csv_path = export_path / "final_rewards.csv"
            final_cols = ["experiment", "bandit_type", "optimize", "eta", "db_episodes",
                         "p1_float_count", "p2_float_count", "data_quality_status",
                         "p1_final", "p1_final_std", "p1_final_se", "p2_final", "p2_final_std", "p2_final_se",
                         "p1_goal", "p1_goal_se", "p2_goal", "p2_goal_se", "avg_final", "avg_goal"]
            if has_ci:
                final_cols.extend(["p1_final_ci", "p2_final_ci", "p1_goal_ci", "p2_goal_ci"])
            available_final_cols = [c for c in final_cols if c in df.columns]
            df[available_final_cols].to_csv(csv_path, index=False)
            console.print(f"[dim]✓ Exported: {csv_path.name}[/]")

    # Sorted Tables (Requested by User)
    if has_final:
        console.print("\n[bold cyan]═══ Sorted Rankings ═══[/]\n")
        
        # Calculate avg_goal and avg_goal_se if possible
        if "p1_goal" in df.columns and "p2_goal" in df.columns:
            df["avg_goal"] = (df["p1_goal"] + df["p2_goal"]) / 2
        else:
            df["avg_goal"] = 0.0

        # Calculate avg_goal_se (average of the SEs)
        if "p1_goal_se" in df.columns and "p2_goal_se" in df.columns:
            df["avg_goal_se"] = (df["p1_goal_se"] + df["p2_goal_se"]) / 2
        else:
            df["avg_goal_se"] = 0.0

        # Helper to format experiment name with baseline
        def format_exp_name(row):
            btype = row['bandit_type']
            name = str(row["experiment"])
            # Simplify name by removing common prefix/suffix if desired, but here we keep full
            # Add baseline tag
            meta = f"({btype})"
            if btype == 'none':
                meta = "[bold red](Baseline)[/]"
            return f"{name} {meta}"

        # 1. Sort by Avg Final
        sorted_final = df.sort_values("avg_final", ascending=False)
        table_final = Table(title="Ranked by Average Final Reward (Descending)")
        table_final.add_column("Rank", justify="right", style="dim")
        table_final.add_column("Experiment", style="cyan")
        table_final.add_column("Avg Final", justify="right", style="bold yellow")
        table_final.add_column("Avg Goal", justify="right")
        table_final.add_column("Avg Goal SE", justify="right", style="dim magenta")

        for i, (_, row) in enumerate(sorted_final.iterrows(), 1):
            table_final.add_row(
                str(i),
                format_exp_name(row),
                f"{row['avg_final']:.4f}",
                f"{row['avg_goal']:.4f}",
                f"{row.get('avg_goal_se', 0):.4f}"
            )
        console.print(table_final)

        # 2. Sort by Avg Goal
        sorted_goal = df.sort_values("avg_goal", ascending=False)
        table_goal = Table(title="Ranked by Average Goal Score (Descending)")
        table_goal.add_column("Rank", justify="right", style="dim")
        table_goal.add_column("Experiment", style="cyan")
        table_goal.add_column("Avg Goal", justify="right", style="bold green")
        table_goal.add_column("Avg Goal SE", justify="right", style="dim magenta")
        table_goal.add_column("Avg Final", justify="right")

        for i, (_, row) in enumerate(sorted_goal.iterrows(), 1):
            table_goal.add_row(
                str(i),
                format_exp_name(row),
                f"{row['avg_goal']:.4f}",
                f"{row.get('avg_goal_se', 0):.4f}",
                f"{row['avg_final']:.4f}"
            )
        console.print(table_goal)

        # 3. Sort by P1 Final Reward
        sorted_p1 = df.sort_values("p1_final", ascending=False)
        table_p1 = Table(title="Ranked by P1 Final Reward (Descending)")
        table_p1.add_column("Rank", justify="right", style="dim")
        table_p1.add_column("Experiment", style="cyan")
        table_p1.add_column("P1 Final", justify="right", style="bold blue")
        table_p1.add_column("P1 Goal", justify="right")
        table_p1.add_column("P1 Goal SE", justify="right", style="dim magenta")

        for i, (_, row) in enumerate(sorted_p1.iterrows(), 1):
            table_p1.add_row(
                str(i),
                format_exp_name(row),
                f"{row['p1_final']:.4f}",
                f"{row['p1_goal']:.4f}",
                f"{row.get('p1_goal_se', 0):.4f}"
            )
        console.print(table_p1)

        # 4. Sort by P2 Final Reward
        sorted_p2 = df.sort_values("p2_final", ascending=False)
        table_p2 = Table(title="Ranked by P2 Final Reward (Descending)")
        table_p2.add_column("Rank", justify="right", style="dim")
        table_p2.add_column("Experiment", style="cyan")
        table_p2.add_column("P2 Final", justify="right", style="bold magenta")
        table_p2.add_column("P2 Goal", justify="right")
        table_p2.add_column("P2 Goal SE", justify="right", style="dim magenta")

        for i, (_, row) in enumerate(sorted_p2.iterrows(), 1):
            table_p2.add_row(
                str(i),
                format_exp_name(row),
                f"{row['p2_final']:.4f}",
                f"{row['p2_goal']:.4f}",
                f"{row.get('p2_goal_se', 0):.4f}"
            )
        console.print(table_p2)

        # 5. Sort by P1 Goal
        sorted_p1_goal = df.sort_values("p1_goal", ascending=False)
        table_p1_goal = Table(title="Ranked by P1 Goal Score (Descending)")
        table_p1_goal.add_column("Rank", justify="right", style="dim")
        table_p1_goal.add_column("Experiment", style="cyan")
        table_p1_goal.add_column("P1 Goal", justify="right", style="bold blue")
        table_p1_goal.add_column("P1 Goal SE", justify="right", style="dim magenta")
        table_p1_goal.add_column("P1 Final", justify="right")

        for i, (_, row) in enumerate(sorted_p1_goal.iterrows(), 1):
            table_p1_goal.add_row(
                str(i),
                format_exp_name(row),
                f"{row['p1_goal']:.4f}",
                f"{row.get('p1_goal_se', 0):.4f}",
                f"{row['p1_final']:.4f}"
            )
        console.print(table_p1_goal)

        # 6. Sort by P2 Goal
        sorted_p2_goal = df.sort_values("p2_goal", ascending=False)
        table_p2_goal = Table(title="Ranked by P2 Goal Score (Descending)")
        table_p2_goal.add_column("Rank", justify="right", style="dim")
        table_p2_goal.add_column("Experiment", style="cyan")
        table_p2_goal.add_column("P2 Goal", justify="right", style="bold magenta")
        table_p2_goal.add_column("P2 Goal SE", justify="right", style="dim magenta")
        table_p2_goal.add_column("P2 Final", justify="right")

        for i, (_, row) in enumerate(sorted_p2_goal.iterrows(), 1):
            table_p2_goal.add_row(
                str(i),
                format_exp_name(row),
                f"{row['p2_goal']:.4f}",
                f"{row.get('p2_goal_se', 0):.4f}",
                f"{row['p2_final']:.4f}"
            )
        console.print(table_p2_goal)

    # Dimension breakdown analysis
    if dim_df is not None and not dim_df.empty:
        display_dimension_analysis(dim_df, use_ci=use_ci, export_dir=export_dir)

    # Summary statistics by bandit type
    if len(df) > 1:
        console.print("\n[bold cyan]═══ Summary by Bandit Type ═══[/]\n")
        agg_cols = {
            "p1_mean": ["mean", "std"],
            "p2_mean": ["mean", "std"],
            "avg": ["mean", "std"],
        }
        if has_final:
            agg_cols["p1_final"] = ["mean", "std"]
            agg_cols["p2_final"] = ["mean", "std"]
            agg_cols["avg_final"] = ["mean", "std"]
        summary = df.groupby("bandit_type").agg(agg_cols).round(4)
        console.print(summary.to_string())

    # Print tags with data quality issues
    quality_cols = ["p1_float_count", "p2_float_count"]
    if all(col in df.columns for col in quality_cols):
        problem_tags = []
        for _, row in df.iterrows():
            p1_float = row.get("p1_float_count", 0)
            p2_float = row.get("p2_float_count", 0)
            if p1_float > 0 or p2_float > 0:
                problem_tags.append(row["experiment"])

        if problem_tags:
            console.print("\n[bold red]═══ Tags with Data Quality Issues ═══[/]\n")
            console.print(f"[yellow]Found {len(problem_tags)} tags with float rewards (missing dimension breakdown):[/]\n")
            for tag in problem_tags:
                console.print(f"[red]{tag}[/]")


def display_dimension_analysis(dim_df: pd.DataFrame, use_ci: bool = False, export_dir: str | None = None) -> None:
    """Display detailed dimension analysis from database."""
    if dim_df.empty:
        return

    title_suffix = " with 95% CI" if use_ci else ""
    console.print(f"\n[bold cyan]═══ Dimension Analysis (from Database){title_suffix} ═══[/]\n")

    # P1 Dimensions Table
    console.print("[bold yellow]P1 (Agent 1) Dimension Scores[/]")
    p1_table = Table(title="P1 Dimensions")
    p1_table.add_column("Experiment", style="cyan")
    p1_table.add_column("Bandit", style="green", justify="center")
    for dim in DIMENSIONS:
        p1_table.add_column(dim[:8], justify="right")  # Truncate dimension names
    p1_table.add_column("Overall", justify="right", style="bold")

    for _, row in dim_df.iterrows():
        # Color the experiment name based on data quality
        exp_name = str(row["experiment"])
        p1_float = row.get("p1_float_count", 0)
        p2_float = row.get("p2_float_count", 0)
        total_eps = row.get("db_episodes", 70)
        
        if p1_float > 0 or p2_float > 0:
            float_pct = ((p1_float + p2_float) / (2 * total_eps) * 100) if total_eps > 0 else 0
            if float_pct >= 20:
                exp_name = f"[bold red]{exp_name}[/]"
            elif float_pct >= 5:
                exp_name = f"[orange]{exp_name}[/]"
            else:
                exp_name = f"[yellow]{exp_name}[/]"
        
        row_data = [exp_name, str(row["bandit_type"])]
        for dim in DIMENSIONS:
            mean_val = row.get(f"p1_{dim}", 0)
            if use_ci:
                ci_val = row.get(f"p1_{dim}_ci", 0)
                row_data.append(f"{mean_val:.2f}±{ci_val:.2f}")
            else:
                std_val = row.get(f"p1_{dim}_std", 0)
                row_data.append(f"{mean_val:.2f}±{std_val:.2f}")
        if use_ci:
            overall_ci = row.get('p1_overall_ci', 0)
            row_data.append(f"{row.get('p1_overall', 0):.2f}±{overall_ci:.2f}")
        else:
            row_data.append(f"{row.get('p1_overall', 0):.2f}±{row.get('p1_overall_std', 0):.2f}")
        p1_table.add_row(*row_data)
    console.print(p1_table)

    # P2 Dimensions Table
    console.print("\n[bold yellow]P2 (Agent 2) Dimension Scores[/]")
    p2_table = Table(title="P2 Dimensions")
    p2_table.add_column("Experiment", style="cyan")
    p2_table.add_column("Bandit", style="green", justify="center")
    for dim in DIMENSIONS:
        p2_table.add_column(dim[:8], justify="right")
    p2_table.add_column("Overall", justify="right", style="bold")

    for _, row in dim_df.iterrows():
        # Color the experiment name based on data quality
        exp_name = str(row["experiment"])
        p1_float = row.get("p1_float_count", 0)
        p2_float = row.get("p2_float_count", 0)
        total_eps = row.get("db_episodes", 70)
        
        if p1_float > 0 or p2_float > 0:
            float_pct = ((p1_float + p2_float) / (2 * total_eps) * 100) if total_eps > 0 else 0
            if float_pct >= 20:
                exp_name = f"[bold red]{exp_name}[/]"
            elif float_pct >= 5:
                exp_name = f"[orange]{exp_name}[/]"
            else:
                exp_name = f"[yellow]{exp_name}[/]"
        
        row_data = [exp_name, str(row["bandit_type"])]
        for dim in DIMENSIONS:
            mean_val = row.get(f"p2_{dim}", 0)
            if use_ci:
                ci_val = row.get(f"p2_{dim}_ci", 0)
                row_data.append(f"{mean_val:.2f}±{ci_val:.2f}")
            else:
                std_val = row.get(f"p2_{dim}_std", 0)
                row_data.append(f"{mean_val:.2f}±{std_val:.2f}")
        if use_ci:
            overall_ci = row.get('p2_overall_ci', 0)
            row_data.append(f"{row.get('p2_overall', 0):.2f}±{overall_ci:.2f}")
        else:
            row_data.append(f"{row.get('p2_overall', 0):.2f}±{row.get('p2_overall_std', 0):.2f}")
        p2_table.add_row(*row_data)
    console.print(p2_table)

    # Combined (P1 + P2) / 2 Table
    console.print("\n[bold yellow]Combined Average (P1 + P2) / 2[/]")
    combined_table = Table(title="Average Dimensions")
    combined_table.add_column("Experiment", style="cyan")
    combined_table.add_column("Bandit", style="green", justify="center")
    for dim in DIMENSIONS:
        combined_table.add_column(dim[:8], justify="right")
    combined_table.add_column("Overall", justify="right", style="bold")

    for _, row in dim_df.iterrows():
        # Color the experiment name based on data quality
        exp_name = str(row["experiment"])
        p1_float = row.get("p1_float_count", 0)
        p2_float = row.get("p2_float_count", 0)
        total_eps = row.get("db_episodes", 70)
        
        if p1_float > 0 or p2_float > 0:
            float_pct = ((p1_float + p2_float) / (2 * total_eps) * 100) if total_eps > 0 else 0
            if float_pct >= 20:
                exp_name = f"[bold red]{exp_name}[/]"
            elif float_pct >= 5:
                exp_name = f"[orange]{exp_name}[/]"
            else:
                exp_name = f"[yellow]{exp_name}[/]"
        
        row_data = [exp_name, str(row["bandit_type"])]
        for dim in DIMENSIONS:
            p1_val = row.get(f"p1_{dim}", 0)
            p2_val = row.get(f"p2_{dim}", 0)
            avg_val = (p1_val + p2_val) / 2
            row_data.append(f"{avg_val:.2f}")
        p1_overall = row.get('p1_overall', 0)
        p2_overall = row.get('p2_overall', 0)
        avg_overall = (p1_overall + p2_overall) / 2
        row_data.append(f"{avg_overall:.2f}")
        combined_table.add_row(*row_data)
    console.print(combined_table)

    # Dimension summary DataFrame
    console.print("\n[bold]Dimension Comparison DataFrame:[/]")
    # Create a summary with experiment and key dimensions
    summary_cols = ["experiment", "bandit_type"]
    for dim in DIMENSIONS:
        summary_cols.append(f"p1_{dim}")
        summary_cols.append(f"p2_{dim}")
    summary_cols.extend(["p1_overall", "p2_overall"])
    available_cols = [c for c in summary_cols if c in dim_df.columns]
    print(dim_df[available_cols].to_string(index=False))
    
    # Export dimension tables to CSV
    if export_dir:
        export_path = Path(export_dir)
        # Export full dimension DataFrame
        csv_path = export_path / "dimension_breakdown.csv"
        dim_df.to_csv(csv_path, index=False)
        console.print(f"\n[dim]✓ Exported: {csv_path.name}[/]")


def _shorten_experiment_name(name: str) -> str:
    """Shorten experiment name for display."""
    if len(name) > 25:
        parts = name.split("_")
        if "adversarial" in name:
            opt_who = "none"
            for p in parts:
                if p in ["none", "p1", "p2", "both"]:
                    opt_who = p
                    break
            return f"opt_{opt_who}"
        else:
            return name[:20] + "..."
    return name


def plot_error_bars(
    df: pd.DataFrame,
    output_path: str | None = None,
    metric: str = "goal",
    show_overlap: bool = True,
    use_numeric_labels: bool = False,
) -> None:
    """
    Plot error bars based on standard error for different experiments.

    This visualizes the mean ± SE for each experiment, making it easy to see
    if confidence intervals overlap (suggesting no significant difference).

    Args:
        df: DataFrame with experiment metrics (from compare_experiments)
        output_path: Path to save the plot (if None, displays interactively)
        metric: Which metric to plot - 'goal', 'final', or 'overall'
        show_overlap: If True, highlight overlapping error bars
        use_numeric_labels: If True, use numeric labels (1, 2, 3...) instead of experiment names
    """
    if df.empty:
        console.print("[red]No data to plot[/]")
        return

    # Determine which columns to use based on metric
    if metric == "goal":
        p1_mean_col = "p1_goal"
        p2_mean_col = "p2_goal"
        p1_se_col = "p1_goal_se"
        p2_se_col = "p2_goal_se"
        title = "Goal Score Comparison with Standard Error"
        ylabel = "Goal Score"
    elif metric == "final":
        p1_mean_col = "p1_final"
        p2_mean_col = "p2_final"
        p1_se_col = "p1_final_se"
        p2_se_col = "p2_final_se"
        title = "Final Reward Comparison with Standard Error"
        ylabel = "Final Reward"
    else:  # overall
        p1_mean_col = "p1_mean"
        p2_mean_col = "p2_mean"
        p1_se_col = "p1_se"
        p2_se_col = "p2_se"
        title = "Per-Turn Reward Comparison with Standard Error"
        ylabel = "Reward"

    # Check if required columns exist
    required_cols = [p1_mean_col, p2_mean_col, p1_se_col, p2_se_col]
    missing_cols = [c for c in required_cols if c not in df.columns]
    if missing_cols:
        console.print(f"[red]Missing columns for plotting: {missing_cols}[/]")
        console.print(f"[dim]Available columns: {list(df.columns)}[/]")
        return

    # Sort by average metric for better visualization
    df = df.copy()
    df["avg_metric"] = (df[p1_mean_col] + df[p2_mean_col]) / 2
    df = df.sort_values("avg_metric", ascending=True).reset_index(drop=True)

    # Experiment names (shortened for display) or numeric labels
    n_exp = len(df)
    x_positions = np.arange(n_exp)

    if use_numeric_labels:
        exp_labels = [str(i + 1) for i in range(n_exp)]
    else:
        exp_labels = [_shorten_experiment_name(name) for name in df["experiment"]]

    # Create figure - vertical error bar plot
    fig, ax = plt.subplots(figsize=(max(10, n_exp * 2), 8))

    # Get means and SEs
    p1_means = df[p1_mean_col].values
    p1_ses = df[p1_se_col].values
    p2_means = df[p2_mean_col].values
    p2_ses = df[p2_se_col].values

    # Offset for side-by-side P1/P2
    offset = 0.15

    # Plot P1 (left dots) with error bars
    ax.errorbar(
        x_positions - offset, p1_means, yerr=p1_ses,
        fmt='o', markersize=10, capsize=8, capthick=2, elinewidth=2,
        color='#3498db', label='Player 1 (Agent)', alpha=0.9
    )

    # Plot P2 (right dots) with error bars
    ax.errorbar(
        x_positions + offset, p2_means, yerr=p2_ses,
        fmt='s', markersize=10, capsize=8, capthick=2, elinewidth=2,
        color='#e74c3c', label='Player 2 (Partner)', alpha=0.9
    )

    # Draw horizontal lines at error bar endpoints for easier comparison
    line_width = 0.08  # Half width of horizontal line
    for i, x in enumerate(x_positions):
        # P1 lines (blue)
        p1_upper = p1_means[i] + p1_ses[i]
        p1_lower = p1_means[i] - p1_ses[i]
        ax.hlines(p1_upper, x - offset - line_width, x - offset + line_width,
                  colors='#3498db', linewidths=2, alpha=0.9)
        ax.hlines(p1_lower, x - offset - line_width, x - offset + line_width,
                  colors='#3498db', linewidths=2, alpha=0.9)
        # P2 lines (red)
        p2_upper = p2_means[i] + p2_ses[i]
        p2_lower = p2_means[i] - p2_ses[i]
        ax.hlines(p2_upper, x + offset - line_width, x + offset + line_width,
                  colors='#e74c3c', linewidths=2, alpha=0.9)
        ax.hlines(p2_lower, x + offset - line_width, x + offset + line_width,
                  colors='#e74c3c', linewidths=2, alpha=0.9)

    # Customize axes
    ax.set_xticks(x_positions)
    ax.set_xticklabels(exp_labels, rotation=0 if use_numeric_labels else 45,
                      ha='center' if use_numeric_labels else 'right', fontsize=11)
    ax.set_xlabel('Experiment ID' if use_numeric_labels else 'Experiment', fontsize=12)
    ax.set_ylabel(ylabel, fontsize=12)
    ax.set_title(title, fontsize=14, fontweight='bold')
    ax.legend(loc='best', fontsize=10)
    ax.grid(axis='y', alpha=0.3, linestyle='--')

    plt.tight_layout()

    # Save or show
    if output_path:
        plt.savefig(output_path, dpi=150, bbox_inches="tight")
        console.print(f"[green]✓ Plot saved to: {output_path}[/]")
    else:
        plt.show()

    plt.close()

    # Print overlap analysis
    if show_overlap:
        _analyze_overlap(df, p1_mean_col, p1_se_col, p2_mean_col, p2_se_col, metric)


def plot_all_metrics_error_bars(
    df: pd.DataFrame,
    output_dir: Path,
    use_numeric_labels: bool = True,
) -> list[str]:
    """
    Plot error bars for all metrics: P1, P2, Avg Final, and Avg Goal.

    Args:
        df: DataFrame with experiment metrics
        output_dir: Directory to save plots
        use_numeric_labels: If True, use numeric labels (1, 2, 3...) instead of experiment names

    Returns a list of saved file paths.
    """
    saved_files = []

    if df.empty:
        console.print("[red]No data to plot[/]")
        return saved_files

    # Sort by avg_final for consistent ordering
    df = df.copy()
    if "avg_final" not in df.columns:
        df["avg_final"] = (df["p1_final"] + df["p2_final"]) / 2
    df = df.sort_values("avg_final", ascending=True).reset_index(drop=True)

    n_exp = len(df)
    x_positions = np.arange(n_exp)

    # Use numeric labels (1, 2, 3...) instead of long experiment names
    if use_numeric_labels:
        exp_labels = [str(i + 1) for i in range(n_exp)]
    else:
        exp_labels = [_shorten_experiment_name(name) for name in df["experiment"]]

    # Define metrics to plot
    metrics_config = [
        {
            "name": "p1_final",
            "mean_col": "p1_final",
            "se_col": "p1_final_se",
            "title": "Player 1 Final Reward",
            "ylabel": "Final Reward",
            "color": "#3498db",
            "filename": "errorbar_p1_final.png",
        },
        {
            "name": "p2_final",
            "mean_col": "p2_final",
            "se_col": "p2_final_se",
            "title": "Player 2 Final Reward",
            "ylabel": "Final Reward",
            "color": "#e74c3c",
            "filename": "errorbar_p2_final.png",
        },
        {
            "name": "avg_final",
            "mean_col": "avg_final",
            "se_col": None,  # Will calculate from p1 and p2
            "title": "Average Final Reward (P1+P2)/2",
            "ylabel": "Avg Final Reward",
            "color": "#2ecc71",
            "filename": "errorbar_avg_final.png",
        },
        {
            "name": "p1_goal",
            "mean_col": "p1_goal",
            "se_col": "p1_goal_se",
            "title": "Player 1 Goal Score",
            "ylabel": "Goal Score",
            "color": "#3498db",
            "filename": "errorbar_p1_goal.png",
        },
        {
            "name": "p2_goal",
            "mean_col": "p2_goal",
            "se_col": "p2_goal_se",
            "title": "Player 2 Goal Score",
            "ylabel": "Goal Score",
            "color": "#e74c3c",
            "filename": "errorbar_p2_goal.png",
        },
        {
            "name": "avg_goal",
            "mean_col": None,  # Will calculate
            "se_col": None,
            "title": "Average Goal Score (P1+P2)/2",
            "ylabel": "Avg Goal Score",
            "color": "#9b59b6",
            "filename": "errorbar_avg_goal.png",
        },
    ]

    # Calculate avg columns if needed
    if "avg_goal" not in df.columns:
        df["avg_goal"] = (df["p1_goal"] + df["p2_goal"]) / 2
    if "avg_goal_se" not in df.columns:
        # SE of average: sqrt((se1^2 + se2^2) / 4)
        df["avg_goal_se"] = np.sqrt((df["p1_goal_se"]**2 + df["p2_goal_se"]**2) / 4)
    if "avg_final_se" not in df.columns:
        df["avg_final_se"] = np.sqrt((df["p1_final_se"]**2 + df["p2_final_se"]**2) / 4)

    # Update config with calculated columns
    for cfg in metrics_config:
        if cfg["name"] == "avg_final":
            cfg["se_col"] = "avg_final_se"
        elif cfg["name"] == "avg_goal":
            cfg["mean_col"] = "avg_goal"
            cfg["se_col"] = "avg_goal_se"

    for cfg in metrics_config:
        mean_col = cfg["mean_col"]
        se_col = cfg["se_col"]

        if mean_col not in df.columns or se_col not in df.columns:
            console.print(f"[yellow]Skipping {cfg['name']}: missing columns[/]")
            continue

        means = df[mean_col].values
        ses = df[se_col].values

        fig, ax = plt.subplots(figsize=(max(10, n_exp * 2), 6))

        # X-axis limits for full-width horizontal lines
        x_min = -0.5
        x_max = n_exp - 0.5

        # Draw full-width horizontal reference lines for overlap comparison
        for i in range(n_exp):
            upper = means[i] + ses[i]
            lower = means[i] - ses[i]
            ax.hlines(upper, x_min, x_max, colors=cfg["color"],
                      linewidths=1.2, alpha=0.3, linestyles='--', zorder=1)
            ax.hlines(lower, x_min, x_max, colors=cfg["color"],
                      linewidths=1.2, alpha=0.3, linestyles='--', zorder=1)

        # Plot error bars with caps
        ax.errorbar(
            x_positions, means, yerr=ses,
            fmt='o', markersize=10, capsize=8, capthick=2, elinewidth=2,
            color=cfg["color"], alpha=0.9, zorder=3
        )

        # Draw short bold horizontal lines at endpoints
        line_width = 0.2
        for i, x in enumerate(x_positions):
            upper = means[i] + ses[i]
            lower = means[i] - ses[i]
            ax.hlines(upper, x - line_width, x + line_width,
                      colors=cfg["color"], linewidths=3, alpha=0.9, zorder=4)
            ax.hlines(lower, x - line_width, x + line_width,
                      colors=cfg["color"], linewidths=3, alpha=0.9, zorder=4)

        ax.set_xlim(x_min, x_max)
        ax.set_xticks(x_positions)
        ax.set_xticklabels(exp_labels, rotation=0 if use_numeric_labels else 45,
                          ha='center' if use_numeric_labels else 'right', fontsize=11)
        ax.set_xlabel('Experiment ID' if use_numeric_labels else 'Experiment', fontsize=12)
        ax.set_ylabel(cfg["ylabel"], fontsize=12)
        ax.set_title(cfg["title"], fontsize=14, fontweight='bold')
        ax.grid(axis='y', alpha=0.3, linestyle='--')

        plt.tight_layout()

        output_path = output_dir / cfg["filename"]
        plt.savefig(output_path, dpi=150, bbox_inches="tight")
        saved_files.append(str(output_path))
        plt.close()

    return saved_files


def plot_summary_all_metrics(
    df: pd.DataFrame,
    output_path: Path,
    use_numeric_labels: bool = True,
) -> str | None:
    """
    Create a comprehensive summary figure combining all 6 error bar plots into
    a single multi-panel figure with subplots (2x3 grid layout).

    Args:
        df: DataFrame with experiment metrics
        output_path: Path to save the summary figure
        use_numeric_labels: If True, use numeric labels (1, 2, 3...) instead of experiment names

    Returns:
        The saved file path, or None if failed.
    """
    if df.empty:
        console.print("[red]No data to plot summary[/]")
        return None

    # Sort by avg_final for consistent ordering
    df = df.copy()
    if "avg_final" not in df.columns:
        df["avg_final"] = (df["p1_final"] + df["p2_final"]) / 2
    df = df.sort_values("avg_final", ascending=True).reset_index(drop=True)

    n_exp = len(df)
    x_positions = np.arange(n_exp)

    # Use numeric labels (1, 2, 3...) instead of long experiment names
    if use_numeric_labels:
        exp_labels = [str(i + 1) for i in range(n_exp)]
    else:
        exp_labels = [_shorten_experiment_name(name) for name in df["experiment"]]

    # Calculate avg columns if needed
    if "avg_goal" not in df.columns:
        df["avg_goal"] = (df["p1_goal"] + df["p2_goal"]) / 2
    if "avg_goal_se" not in df.columns:
        df["avg_goal_se"] = np.sqrt((df["p1_goal_se"]**2 + df["p2_goal_se"]**2) / 4)
    if "avg_final_se" not in df.columns:
        df["avg_final_se"] = np.sqrt((df["p1_final_se"]**2 + df["p2_final_se"]**2) / 4)

    # Define metrics to plot in 2x3 grid (row-major order)
    metrics_config = [
        # Row 1: Final Rewards
        {
            "mean_col": "p1_final",
            "se_col": "p1_final_se",
            "title": "Player 1 Final Reward",
            "ylabel": "Final Reward",
            "color": "#3498db",
        },
        {
            "mean_col": "p2_final",
            "se_col": "p2_final_se",
            "title": "Player 2 Final Reward",
            "ylabel": "Final Reward",
            "color": "#e74c3c",
        },
        {
            "mean_col": "avg_final",
            "se_col": "avg_final_se",
            "title": "Average Final Reward",
            "ylabel": "Avg Final Reward",
            "color": "#2ecc71",
        },
        # Row 2: Goal Scores
        {
            "mean_col": "p1_goal",
            "se_col": "p1_goal_se",
            "title": "Player 1 Goal Score",
            "ylabel": "Goal Score",
            "color": "#3498db",
        },
        {
            "mean_col": "p2_goal",
            "se_col": "p2_goal_se",
            "title": "Player 2 Goal Score",
            "ylabel": "Goal Score",
            "color": "#e74c3c",
        },
        {
            "mean_col": "avg_goal",
            "se_col": "avg_goal_se",
            "title": "Average Goal Score",
            "ylabel": "Avg Goal Score",
            "color": "#9b59b6",
        },
    ]

    # Create 2x3 grid figure
    fig, axes = plt.subplots(2, 3, figsize=(20, 12))
    axes = axes.flatten()

    # X-axis limits for full-width horizontal lines
    x_min = -0.5
    x_max = n_exp - 0.5

    for idx, cfg in enumerate(metrics_config):
        ax = axes[idx]
        mean_col = cfg["mean_col"]
        se_col = cfg["se_col"]
        color = cfg["color"]

        if mean_col not in df.columns or se_col not in df.columns:
            ax.text(0.5, 0.5, f"Missing: {mean_col}", ha='center', va='center',
                    transform=ax.transAxes, fontsize=12, color='red')
            ax.set_title(cfg["title"], fontsize=12, fontweight='bold')
            continue

        means = df[mean_col].values
        ses = df[se_col].values

        # First draw horizontal reference lines spanning full width for overlap comparison
        # Use lighter colors for the extended lines
        line_colors = ['#a8d5e5', '#f5b7b1', '#a9dfbf', '#a8d5e5', '#f5b7b1', '#d7bde2']
        light_color = line_colors[idx]

        for i in range(n_exp):
            upper = means[i] + ses[i]
            lower = means[i] - ses[i]

            # Draw full-width horizontal lines at upper and lower bounds
            ax.hlines(upper, x_min, x_max, colors=light_color,
                      linewidths=1.5, alpha=0.6, linestyles='--', zorder=1)
            ax.hlines(lower, x_min, x_max, colors=light_color,
                      linewidths=1.5, alpha=0.6, linestyles='--', zorder=1)

        # Plot error bars with caps on top
        ax.errorbar(
            x_positions, means, yerr=ses,
            fmt='o', markersize=10, capsize=8, capthick=2, elinewidth=2,
            color=color, alpha=0.9, zorder=3
        )

        # Draw short horizontal lines at endpoints (bold, to mark the exact bounds)
        line_extend = 0.2
        for i, x in enumerate(x_positions):
            upper = means[i] + ses[i]
            lower = means[i] - ses[i]
            ax.hlines(upper, x - line_extend, x + line_extend,
                      colors=color, linewidths=3, alpha=0.9, zorder=4)
            ax.hlines(lower, x - line_extend, x + line_extend,
                      colors=color, linewidths=3, alpha=0.9, zorder=4)

        # Customize subplot
        ax.set_xticks(x_positions)
        ax.set_xticklabels(exp_labels, rotation=0 if use_numeric_labels else 45,
                          ha='center' if use_numeric_labels else 'right', fontsize=9)
        ax.set_ylabel(cfg["ylabel"], fontsize=11)
        ax.set_title(cfg["title"], fontsize=12, fontweight='bold')
        ax.grid(axis='y', alpha=0.3, linestyle='--')

        # Add minor gridlines for better precision reading
        ax.yaxis.set_minor_locator(plt.MultipleLocator(0.05))
        ax.grid(axis='y', which='minor', alpha=0.15, linestyle=':')

    # Add overall title
    fig.suptitle('Error Bar Summary: All Metrics Comparison\n(Mean ± Standard Error)',
                 fontsize=16, fontweight='bold', y=1.02)

    # Adjust layout
    plt.tight_layout()
    plt.subplots_adjust(top=0.92, hspace=0.35, wspace=0.25)

    # Save figure
    plt.savefig(output_path, dpi=150, bbox_inches="tight", facecolor='white')
    plt.close()

    return str(output_path)


def save_comparison_results(
    df: pd.DataFrame,
    dim_df: pd.DataFrame,
    experiments: list[str],
    output_dir: Path | None = None,
) -> Path:
    """
    Save comparison results to a structured output directory.

    Creates:
    - metadata.json: experiment info and timestamp (with experiment_id_map)
    - results.csv: main comparison results
    - dimensions.csv: dimension breakdown
    - errorbar_*.png: error bar plots for all metrics
    - overlap_analysis.txt: overlap analysis text

    Returns the output directory path.
    """
    import hashlib
    from datetime import datetime

    # Generate unique directory name based on experiment tags (not timestamp)
    # Same tags will always go to the same directory
    if output_dir is None:
        # Create hash from sorted experiment names for consistency
        exp_str = "_".join(sorted(experiments))
        hash_suffix = hashlib.md5(exp_str.encode()).hexdigest()[:8]
        dir_name = f"comparison_{hash_suffix}"
        output_dir = OUTPUTS_DIR / dir_name

    output_dir.mkdir(parents=True, exist_ok=True)

    # Create experiment ID mapping (1, 2, 3, ... for each experiment)
    # Sort by avg_final for consistent ordering with plots
    df_sorted = df.copy()
    if "avg_final" not in df_sorted.columns:
        df_sorted["avg_final"] = (df_sorted["p1_final"] + df_sorted["p2_final"]) / 2
    df_sorted = df_sorted.sort_values("avg_final", ascending=True).reset_index(drop=True)

    exp_id_map = {}
    for i, (_, row) in enumerate(df_sorted.iterrows(), 1):
        exp_name = str(row["experiment"])
        exp_id_map[str(i)] = exp_name  # 1-indexed

    # Save metadata
    metadata = {
        "experiments": experiments,
        "n_experiments": len(experiments),
        "timestamp": datetime.now().isoformat(),
        "metrics_calculated": list(df.columns),
        "experiment_id_map": exp_id_map,
    }
    with open(output_dir / "metadata.json", "w") as f:
        json.dump(metadata, f, indent=2)

    # Save main results
    df.to_csv(output_dir / "results.csv", index=False)

    # Save dimension breakdown
    if not dim_df.empty:
        dim_df.to_csv(output_dir / "dimensions.csv", index=False)

    # Generate all error bar plots (with numeric labels for cleaner display)
    plot_files = plot_all_metrics_error_bars(df, output_dir, use_numeric_labels=True)

    # Generate combined P1/P2 comparison plots (with numeric labels)
    for metric in ["goal", "final"]:
        plot_error_bars(
            df,
            output_path=str(output_dir / f"errorbar_comparison_{metric}.png"),
            metric=metric,
            show_overlap=False,  # Don't print to console here
            use_numeric_labels=True,
        )
        plot_files.append(str(output_dir / f"errorbar_comparison_{metric}.png"))

    # Generate comprehensive summary figure with all 6 metrics (with numeric labels)
    summary_path = output_dir / "errorbar_summary_all.png"
    summary_file = plot_summary_all_metrics(df, summary_path, use_numeric_labels=True)
    if summary_file:
        plot_files.append(summary_file)
        console.print(f"[green]✓ Summary plot saved to: {summary_path}[/]")

    # Save overlap analysis to file
    overlap_text = _get_overlap_analysis_text(df)
    with open(output_dir / "overlap_analysis.txt", "w") as f:
        f.write(overlap_text)

    console.print(f"\n[green]═══ Results Saved to Directory ═══[/]")
    console.print(f"[cyan]Directory: {output_dir}[/]")
    console.print(f"[dim]  - metadata.json[/]")
    console.print(f"[dim]  - results.csv[/]")
    if not dim_df.empty:
        console.print(f"[dim]  - dimensions.csv[/]")
    console.print(f"[dim]  - overlap_analysis.txt[/]")
    for pf in plot_files:
        console.print(f"[dim]  - {Path(pf).name}[/]")

    return output_dir


def _get_overlap_analysis_text(df: pd.DataFrame) -> str:
    """Generate overlap analysis text for all metrics."""
    lines = []
    lines.append("=" * 60)
    lines.append("OVERLAP ANALYSIS REPORT")
    lines.append("=" * 60)
    lines.append("")

    metrics = [
        ("Goal Score", "p1_goal", "p1_goal_se", "p2_goal", "p2_goal_se"),
        ("Final Reward", "p1_final", "p1_final_se", "p2_final", "p2_final_se"),
    ]

    for metric_name, p1_mean, p1_se, p2_mean, p2_se in metrics:
        if p1_mean not in df.columns:
            continue

        lines.append(f"\n{'-' * 40}")
        lines.append(f"Metric: {metric_name}")
        lines.append(f"{'-' * 40}")

        # Calculate averages
        df_copy = df.copy()
        df_copy["avg_mean"] = (df_copy[p1_mean] + df_copy[p2_mean]) / 2
        df_copy["avg_se"] = (df_copy[p1_se] + df_copy[p2_se]) / 2

        experiments = df_copy["experiment"].tolist()
        n = len(experiments)

        sig_diff = []
        overlapping = []

        for i in range(n):
            for j in range(i + 1, n):
                exp1 = experiments[i][:30]
                exp2 = experiments[j][:30]
                mean1 = df_copy["avg_mean"].iloc[i]
                mean2 = df_copy["avg_mean"].iloc[j]
                se1 = df_copy["avg_se"].iloc[i]
                se2 = df_copy["avg_se"].iloc[j]

                upper1 = mean1 + se1
                lower1 = mean1 - se1
                upper2 = mean2 + se2
                lower2 = mean2 - se2

                if upper1 < lower2 or upper2 < lower1:
                    gap = min(abs(lower2 - upper1), abs(lower1 - upper2))
                    sig_diff.append((exp1, mean1, exp2, mean2, gap))
                else:
                    overlapping.append((exp1, exp2))

        lines.append(f"\nSignificant Differences (No Overlap): {len(sig_diff)}")
        for exp1, m1, exp2, m2, gap in sig_diff:
            lines.append(f"  {exp1} ({m1:.4f}) vs {exp2} ({m2:.4f}) - Gap: {gap:.4f}")

        lines.append(f"\nOverlapping (Not Significant): {len(overlapping)}")
        for exp1, exp2 in overlapping:
            lines.append(f"  {exp1} vs {exp2}")

        total_pairs = n * (n - 1) // 2
        lines.append(f"\nSummary: {len(sig_diff)}/{total_pairs} pairs show significant difference")

    return "\n".join(lines)


def _analyze_overlap(
    df: pd.DataFrame,
    p1_mean_col: str,
    p1_se_col: str,
    p2_mean_col: str,
    p2_se_col: str,
    metric: str,
) -> None:
    """Analyze and report which experiments have overlapping error bars."""
    console.print(f"\n[bold cyan]═══ Overlap Analysis ({metric.upper()}) ═══[/]\n")

    experiments = df["experiment"].tolist()
    n = len(experiments)

    if n < 2:
        console.print("[dim]Need at least 2 experiments to analyze overlap[/]")
        return

    # Calculate avg metric and SE for comparison
    df = df.copy()
    df["avg_mean"] = (df[p1_mean_col] + df[p2_mean_col]) / 2
    df["avg_se"] = (df[p1_se_col] + df[p2_se_col]) / 2

    # Check pairwise overlaps
    overlaps = []
    non_overlaps = []

    for i in range(n):
        for j in range(i + 1, n):
            exp1, exp2 = experiments[i], experiments[j]
            mean1, se1 = df.iloc[i]["avg_mean"], df.iloc[i]["avg_se"]
            mean2, se2 = df.iloc[j]["avg_mean"], df.iloc[j]["avg_se"]

            # Check if error bars overlap (mean ± SE)
            lower1, upper1 = mean1 - se1, mean1 + se1
            lower2, upper2 = mean2 - se2, mean2 + se2

            # Overlap if ranges intersect
            has_overlap = not (upper1 < lower2 or upper2 < lower1)

            # Calculate gap (negative = overlap, positive = separation)
            gap = max(lower1, lower2) - min(upper1, upper2)

            if has_overlap:
                overlaps.append((exp1, exp2, mean1, mean2, gap))
            else:
                non_overlaps.append((exp1, exp2, mean1, mean2, gap))

    # Display results
    if non_overlaps:
        console.print("[bold green]Significant Differences (No Overlap):[/]")
        table = Table(show_header=True)
        table.add_column("Experiment 1", style="cyan")
        table.add_column("Mean 1", justify="right")
        table.add_column("Experiment 2", style="cyan")
        table.add_column("Mean 2", justify="right")
        table.add_column("Gap", justify="right", style="green")

        for exp1, exp2, m1, m2, gap in sorted(non_overlaps, key=lambda x: -x[4])[:10]:
            table.add_row(
                exp1[:25], f"{m1:.4f}",
                exp2[:25], f"{m2:.4f}",
                f"+{gap:.4f}"
            )
        console.print(table)
        if len(non_overlaps) > 10:
            console.print(f"[dim]... and {len(non_overlaps) - 10} more pairs[/]")

    if overlaps:
        console.print(f"\n[bold yellow]Overlapping (Not Significantly Different): {len(overlaps)} pairs[/]")
        if len(overlaps) <= 10:
            for exp1, exp2, m1, m2, gap in overlaps:
                console.print(f"  [dim]• {exp1[:30]} vs {exp2[:30]}[/]")

    # Summary
    total_pairs = n * (n - 1) // 2
    console.print(f"\n[bold]Summary:[/] {len(non_overlaps)}/{total_pairs} pairs show significant difference")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Evaluate experiments from outputs directory",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # List all experiments
  python evaluate_by_tag.py --list-tags

  # Evaluate a single experiment
  python evaluate_by_tag.py --tag bandit_adversarial_p1_20251204_181158

  # Compare multiple experiments
  python evaluate_by_tag.py --tags exp1 exp2 exp3

  # Compare all experiments with a pattern
  python evaluate_by_tag.py --pattern "bandit_adversarial"

  # Export results to CSV
  python evaluate_by_tag.py --tags exp1 exp2 --output comparison.csv
"""
    )
    parser.add_argument(
        "--tag",
        type=str,
        help="Single experiment folder name to evaluate",
    )
    parser.add_argument(
        "--tags",
        type=str,
        nargs="+",
        help="Multiple experiment folder names to compare",
    )
    parser.add_argument(
        "--pattern",
        type=str,
        help="Pattern to match experiment folder names (e.g., 'bandit_adversarial')",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Output file for results (JSON or CSV based on extension)",
    )
    parser.add_argument(
        "--list-tags",
        action="store_true",
        help="List all available experiments in the outputs directory",
    )
    parser.add_argument(
        "--outputs-dir",
        type=str,
        default=None,
        help="Custom outputs directory path",
    )
    parser.add_argument(
        "--use-ci",
        action="store_true",
        help="Display 95%% confidence intervals for final rewards (requires database access)",
    )
    parser.add_argument(
        "--export-csv",
        type=str,
        default=None,
        help="Export all tables to CSV files in the specified directory",
    )
    parser.add_argument(
        "--plot",
        type=str,
        default=None,
        metavar="PATH",
        help="Generate error bar plot and save to the specified path (e.g., plot.png)",
    )
    parser.add_argument(
        "--plot-metric",
        type=str,
        choices=["goal", "final", "overall"],
        default="goal",
        help="Which metric to plot: goal, final, or overall (default: goal)",
    )
    parser.add_argument(
        "--save-all",
        action="store_true",
        help="Save all results, plots, and metadata to a structured output directory",
    )
    parser.add_argument(
        "--csv",
        type=str,
        default=None,
        help="Read experiment tags from a CSV file (expects a 'tag' column)",
    )
    parser.add_argument(
        "--output-xlsx",
        type=str,
        default=None,
        help="Export the summary table to an XLSX file",
    )
    parser.add_argument(
        "--parallel",
        type=int,
        default=8,
        help="Number of parallel workers for loading experiment data (default: 8)",
    )
    parser.add_argument(
        "--eval-set",
        type=str,
        nargs="+",
        choices=["all", "hard"],
        default=None,
        help="Filter evaluation by dataset: 'all' for complete dataset (90 envs), 'hard' for difficult subset (14 envs). "
             "Can specify multiple values (e.g., --eval-set all hard) to generate a side-by-side comparison table.",
    )
    return parser.parse_args()



def list_available_experiments(outputs_dir: Path = OUTPUTS_DIR) -> None:
    """List available experiments by scanning outputs directory."""
    if not outputs_dir.exists():
        console.print(f"[red]Outputs directory not found: {outputs_dir}[/]")
        return

    console.print(f"[cyan]Scanning outputs directory: {outputs_dir}[/]\n")

    # Scan experiment folders
    experiments = []
    for exp_dir in sorted(outputs_dir.iterdir(), reverse=True):
        if not exp_dir.is_dir():
            continue

        results_file = exp_dir / "results.json"
        batch_results_file = exp_dir / "batch_results.json"
        
        has_results = results_file.exists() or batch_results_file.exists()

        exp_info = {
            "name": exp_dir.name,
            "has_results": has_results,
            "bandit_type": "unknown",
            "optimize": "unknown",
            "turns": 0,
            "p1_mean": 0.0,
        }

        if has_results:
            try:
                # Load using the unified loader to handle both formats
                data = load_experiment_results(exp_dir.name, outputs_dir)
                if data:
                    exp_info["bandit_type"] = data.get("bandit_type", "unknown")
                    exp_info["optimize"] = data.get("optimize_mode", "unknown")
                    exp_info["turns"] = data.get("total_turns", 0)
                    exp_info["p1_mean"] = data.get("p1_avg_reward", 0.0)
            except Exception:
                pass

        experiments.append(exp_info)

    if not experiments:
        console.print("[yellow]No experiments found. Run some experiments first.[/]")
        return

    # Count statistics
    with_results = sum(1 for e in experiments if e["has_results"])

    table = Table(title=f"Available Experiments ({len(experiments)} total, {with_results} with results)")
    table.add_column("Experiment Name", style="cyan", max_width=45)
    table.add_column("Results", justify="center")
    table.add_column("Bandit", justify="center", style="green")
    table.add_column("Opt", justify="center")
    table.add_column("Turns", justify="right")
    table.add_column("P1 Mean", justify="right")

    for exp in experiments[:40]:  # Show up to 40 experiments
        status = "[green]✓[/]" if exp["has_results"] else "[red]✗[/]"
        p1_mean = f"{exp['p1_mean']:.2f}" if exp["has_results"] else "-"
        turns = str(exp["turns"]) if exp["has_results"] else "-"

        table.add_row(
            exp["name"][-45:],
            status,
            exp["bandit_type"] if exp["has_results"] else "-",
            exp["optimize"] if exp["has_results"] else "-",
            turns,
            p1_mean,
        )

    console.print(table)

    if len(experiments) > 40:
        console.print(f"\n[dim]... and {len(experiments) - 40} more experiments[/]")

    console.print("\n[dim]Tip: Use --tag <experiment_name> to evaluate a specific experiment[/]")
    console.print("[dim]     Use --pattern <pattern> to compare experiments matching a pattern[/]")


def get_experiments_by_pattern(pattern: str, outputs_dir: Path = OUTPUTS_DIR) -> list[str]:
    """Get experiment names matching a pattern."""
    experiments = []
    for exp_dir in outputs_dir.iterdir():
        if exp_dir.is_dir() and pattern in exp_dir.name:
            results_file = exp_dir / "results.json"
            if results_file.exists():
                experiments.append(exp_dir.name)
    return sorted(experiments)


def main() -> None:
    args = parse_args()

    # Determine outputs directory
    outputs_dir = Path(args.outputs_dir) if args.outputs_dir else OUTPUTS_DIR

    if args.list_tags:
        list_available_experiments(outputs_dir)
        return

    # Load environment IDs filter(s) if --eval-set is specified
    eval_sets: list[str] = args.eval_set if args.eval_set else []
    multi_evalset_mode = len(eval_sets) > 1

    # For single eval-set mode, prepare the filter
    env_ids_filter: set[str] | None = None
    if len(eval_sets) == 1:
        env_ids_filter = load_env_ids_by_dataset(eval_sets[0])
        if env_ids_filter:
            console.print(f"[cyan]Filtering by eval-set '{eval_sets[0]}': {len(env_ids_filter)} environment IDs[/]\n")
        elif eval_sets[0] == "hard":
            console.print(f"[yellow]⚠ Warning: Could not load environment IDs for '{eval_sets[0]}' dataset[/]\n")
    elif multi_evalset_mode:
        console.print(f"[cyan]Multi-dataset comparison mode: {eval_sets}[/]\n")

    # Handle pattern matching
    experiments = []
    if args.csv:
        # Read tags from CSV file
        csv_path = Path(args.csv)
        if not csv_path.exists():
            console.print(f"[red]Error: CSV file not found: {args.csv}[/]")
            return
        try:
            # Read CSV without header - each line is a tag
            csv_df = pd.read_csv(csv_path, header=None)
            experiments = [str(x).strip() for x in csv_df.iloc[:, 0].dropna().tolist() if str(x).strip()]
            console.print(f"[green]Loaded {len(experiments)} experiments from CSV: {args.csv}[/]\n")
        except Exception as e:
            console.print(f"[red]Error reading CSV file: {e}[/]")
            return
    elif args.pattern:
        experiments = get_experiments_by_pattern(args.pattern, outputs_dir)
        if not experiments:
            console.print(f"[red]No experiments found matching pattern: {args.pattern}[/]")
            return
        console.print(f"[green]Found {len(experiments)} experiments matching '{args.pattern}'[/]\n")
    elif args.tags:
        experiments = args.tags
    elif args.tag:
        experiments = [args.tag]
    else:
        console.print("[red]Error: Please specify --tag, --tags, --pattern, --csv, or --list-tags[/]")
        return

    if len(experiments) == 1:
        # Single experiment evaluation
        metrics = calculate_metrics_for_experiment(experiments[0], outputs_dir, env_ids_filter=env_ids_filter)
        display_metrics(metrics)

        if args.output_xlsx:
            summary_evalset = eval_sets[0] if eval_sets else "default"
            df, dim_df = compare_experiments(
                experiments,
                outputs_dir,
                use_ci=args.use_ci,
                max_workers=args.parallel,
                env_ids_filter=env_ids_filter,
            )
            summary_df = build_summary_dataframe({summary_evalset: (df, dim_df)})
            if summary_df.empty:
                console.print("[red]No summary data available to export[/]")
                return
            try:
                export_multi_evalset_summary_xlsx(summary_df, args.output_xlsx)
            except RuntimeError as exc:
                console.print(f"[red]Error: {exc}[/]")
                return
            console.print(f"\n[green]XLSX summary saved to: {args.output_xlsx}[/]")

        if args.output:
            with open(args.output, "w") as f:
                json.dump(metrics, f, indent=2)
            console.print(f"\n[green]Results saved to: {args.output}[/]")
    else:
        # Multiple experiments comparison
        if multi_evalset_mode:
            # Multi eval-set mode: run comparison for each eval-set and display side-by-side
            results_by_evalset: dict[str, tuple[pd.DataFrame, pd.DataFrame]] = {}
            for evalset in eval_sets:
                console.print(f"\n[bold cyan]Loading data for eval-set: {evalset}[/]")
                evalset_filter = load_env_ids_by_dataset(evalset)
                df, dim_df = compare_experiments(
                    experiments, outputs_dir, use_ci=args.use_ci,
                    max_workers=args.parallel, env_ids_filter=evalset_filter
                )
                results_by_evalset[evalset] = (df, dim_df)

            # Display the multi-dataset comparison table
            display_multi_evalset_comparison(results_by_evalset, use_ci=args.use_ci)

            if args.output_xlsx:
                summary_df = build_summary_dataframe(results_by_evalset)
                if summary_df.empty:
                    console.print("[red]No multi-dataset summary data available to export[/]")
                else:
                    try:
                        export_multi_evalset_summary_xlsx(summary_df, args.output_xlsx)
                    except RuntimeError as exc:
                        console.print(f"[red]Error: {exc}[/]")
                        return
                    console.print(f"\n[green]XLSX summary saved to: {args.output_xlsx}[/]")

            # Also display individual tables if requested
            if args.export_csv:
                for evalset, (df, dim_df) in results_by_evalset.items():
                    export_subdir = f"{args.export_csv}/{evalset}"
                    display_comparison(df, dim_df, use_ci=args.use_ci, export_dir=export_subdir)
        else:
            # Single eval-set mode (or no eval-set specified)
            df, dim_df = compare_experiments(
                experiments, outputs_dir, use_ci=args.use_ci, max_workers=args.parallel, env_ids_filter=env_ids_filter
            )
            display_comparison(df, dim_df, use_ci=args.use_ci, export_dir=args.export_csv)

            if args.output_xlsx:
                summary_evalset = eval_sets[0] if eval_sets else "default"
                summary_df = build_summary_dataframe({summary_evalset: (df, dim_df)})
                if summary_df.empty:
                    console.print("[red]No summary data available to export[/]")
                else:
                    try:
                        export_multi_evalset_summary_xlsx(summary_df, args.output_xlsx)
                    except RuntimeError as exc:
                        console.print(f"[red]Error: {exc}[/]")
                        return
                    console.print(f"\n[green]XLSX summary saved to: {args.output_xlsx}[/]")

            # Save all results to structured directory if requested
            if args.save_all:
                save_comparison_results(df, dim_df, experiments)

            # Generate error bar plot if requested
            if args.plot:
                plot_error_bars(df, output_path=args.plot, metric=args.plot_metric)

            if args.output:
                if args.output.endswith(".csv"):
                    df.to_csv(args.output, index=False)
                    # Also save dimension data
                    if not dim_df.empty:
                        dim_output = args.output.replace(".csv", "_dimensions.csv")
                        dim_df.to_csv(dim_output, index=False)
                        console.print(f"[green]Dimension data saved to: {dim_output}[/]")
                else:
                    df.to_json(args.output, orient="records", indent=2)
                console.print(f"\n[green]Results saved to: {args.output}[/]")


if __name__ == "__main__":
    main()
